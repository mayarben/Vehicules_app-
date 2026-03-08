# utils/merge_global.py
# -*- coding: utf-8 -*-

import re
from io import BytesIO
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# CONFIG (same as your script)
# =========================
priority_sheets = [
    "TAS Vehicle List",
    "Peugeot Vehicle List",
    "Citroen Vehicle List"
]

vehicle_list_sheets = {
    "TAS": "TAS Vehicle List",
    "Peugeot": "Peugeot Vehicle List",
    "Citroen": "Citroen Vehicle List",
}

vehicle_pattern = re.compile(r"^\d{2}-\d{6}$")

def infer_brand_from_name(name: str) -> str:
    p = (name or "").lower()
    if "tas" in p:
        return "TAS"
    if "peugeot" in p:
        return "Peugeot"
    if "citroen" in p or "citreon" in p:
        return "Citroen"
    return "Unknown"

# =========================
# STYLE HELPERS (yours)
# =========================
THIN = Side(style="thin", color="FF000000")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_LIST_HEADER = PatternFill("solid", fgColor="FF8B2E2E")
FILL_MAIN_HEADER = PatternFill("solid", fgColor="FFF4C7A1")
FILL_PIECE_HEADER = PatternFill("solid", fgColor="FFCFE2F3")
FILL_TOTAL = PatternFill("solid", fgColor="FFC6E0B4")

FONT_WHITE_BOLD = Font(bold=True, color="FFFFFFFF")
FONT_BOLD = Font(bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def _apply_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = BORDER_THIN

def _normalize(s):
    return str(s).strip().lower().replace("\u00a0", " ")

def style_vehicle_list(ws):
    max_row = ws.max_row
    max_col = ws.max_column

    header_row = None
    for r in range(1, min(15, max_row) + 1):
        row_vals = [_normalize(ws.cell(r, c).value) for c in range(1, min(10, max_col) + 1)]
        if any("vehicule id" in v or "véhicule id" in v for v in row_vals):
            header_row = r
            break
    if header_row is None:
        return

    table_max_col = min(max_col, 3) if max_col >= 3 else max_col

    for c in range(1, table_max_col + 1):
        cell = ws.cell(header_row, c)
        cell.fill = FILL_LIST_HEADER
        cell.font = FONT_WHITE_BOLD
        cell.alignment = ALIGN_CENTER

    last = header_row
    for r in range(header_row + 1, max_row + 1):
        if ws.cell(r, 1).value is None and ws.cell(r, 2).value is None and ws.cell(r, 3).value is None:
            if r > header_row + 2:
                break
        else:
            last = r

    _apply_border_range(ws, header_row, last, 1, table_max_col)

    for r in range(header_row + 1, last + 1):
        ws.cell(r, 1).font = FONT_BOLD
        ws.cell(r, 1).alignment = ALIGN_LEFT
        ws.cell(r, 2).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(r, 3).alignment = ALIGN_LEFT

def _find_row_contains(ws, text, col=1, search_rows=200):
    t = _normalize(text)
    for r in range(1, min(search_rows, ws.max_row) + 1):
        v = ws.cell(r, col).value
        if v is None:
            continue
        if t in _normalize(v):
            return r
    return None

def _style_section(ws, title_row, header_fill):
    if title_row is None:
        return

    header_row = title_row + 1
    max_col = ws.max_column

    last_col = 0
    for c in range(1, max_col + 1):
        if ws.cell(header_row, c).value is not None:
            last_col = c
    if last_col == 0:
        return

    for c in range(1, last_col + 1):
        cell = ws.cell(header_row, c)
        cell.fill = header_fill
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    end_row = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is not None:
            an = _normalize(a)
            if "piece" in an or "pièce" in an or "main d'oeuvre" in an or "main d'œuvre" in an:
                break

        if any(ws.cell(r, c).value is not None for c in range(1, last_col + 1)):
            end_row = r
        else:
            if r > header_row + 1:
                break

    _apply_border_range(ws, header_row, end_row, 1, last_col)

    for r in range(header_row + 1, end_row + 1):
        for c in range(1, last_col + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                ws.cell(r, c).alignment = Alignment(horizontal="right", vertical="center")
            else:
                ws.cell(r, c).alignment = ALIGN_LEFT

def style_vehicle_detail(ws):
    main_row = _find_row_contains(ws, "main d", col=1, search_rows=200)
    piece_row = _find_row_contains(ws, "pièce", col=1, search_rows=400) or _find_row_contains(ws, "piece", col=1, search_rows=400)

    if main_row:
        ws.cell(main_row, 1).font = Font(size=14, bold=True)
    if piece_row:
        ws.cell(piece_row, 1).font = Font(size=14, bold=True)

    _style_section(ws, main_row, FILL_MAIN_HEADER)
    _style_section(ws, piece_row, FILL_PIECE_HEADER)

    total_row = _find_row_contains(ws, "total htva", col=1, search_rows=800) or _find_row_contains(ws, "total htva", col=2, search_rows=800)
    if total_row:
        val_col = None
        for c in range(2, min(ws.max_column, 12) + 1):
            v = ws.cell(total_row, c).value
            if isinstance(v, (int, float)):
                val_col = c
                break

        ws.cell(total_row, 1).font = Font(size=12, bold=True)
        ws.cell(total_row, 1).alignment = ALIGN_LEFT

        if val_col:
            vc = ws.cell(total_row, val_col)
            vc.fill = FILL_TOTAL
            vc.font = Font(bold=True)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = BORDER_THIN

# =========================
# Copy sheet exactly (yours)
# =========================
def copy_sheet(source_sheet, master_wb, new_title):
    new_sheet = master_wb.create_sheet(title=new_title)

    for col, dim in source_sheet.column_dimensions.items():
        new_sheet.column_dimensions[col] = copy(dim)

    for row, dim in source_sheet.row_dimensions.items():
        new_sheet.row_dimensions[row] = copy(dim)

    for merged in source_sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged))

    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.alignment = copy(cell.alignment)
                new_cell.border = copy(cell.border)

            if cell.hyperlink:
                new_cell.hyperlink = copy(cell.hyperlink)

            if cell.number_format:
                new_cell.number_format = cell.number_format

    if source_sheet.freeze_panes:
        new_sheet.freeze_panes = source_sheet.freeze_panes

    new_sheet.sheet_properties.tabColor = source_sheet.sheet_properties.tabColor

    if "vehicle list" in (new_title or "").lower():
        style_vehicle_list(new_sheet)
    else:
        style_vehicle_detail(new_sheet)

    return new_sheet

# =========================
# GLOBAL SHEET STYLES (yours)
# =========================
thin2 = Side(style="thin", color="1F1F1F")
border2 = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F4E79")

section_font = Font(bold=True, color="FFFFFF", size=12)
section_fill = PatternFill("solid", fgColor="2F5597")

subheader_font = Font(bold=True, color="FFFFFF")
subheader_fill = PatternFill("solid", fgColor="1F4E79")

label_font = Font(bold=True)

wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

def set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

def write_row(ws, row_idx, values, start_col=1, *,
              font=None, fill=None, alignment=None, apply_border=False):
    for j, v in enumerate(values, start=start_col):
        c = ws.cell(row_idx, j, v)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        if alignment:
            c.alignment = alignment
        if apply_border:
            c.border = border2

# =========================
# Vehicle list reader (yours)
# =========================
def find_header_row_and_cols(ws):
    max_row = min(ws.max_row, 50)
    max_col = min(ws.max_column, 30)
    for r in range(1, max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        norm = [_normalize(v) for v in vals]
        if any("véhicule id" in v or "vehicule id" in v for v in norm):
            col_vid = None
            col_total = None
            col_link = None
            for c, v in enumerate(norm, start=1):
                if col_vid is None and ("véhicule id" in v or "vehicule id" in v):
                    col_vid = c
                if col_total is None and ("total htva" in v):
                    col_total = c
                if col_link is None and ("open sheet" in v or "open" == v):
                    col_link = c
            if col_vid is None:
                col_vid = 1
            if col_total is None:
                col_total = 2
            if col_link is None:
                col_link = 3 if ws.max_column >= 3 else None
            return r, col_vid, col_total, col_link
    return None, None, None, None

def read_vehicle_list(ws, brand):
    header_row, col_vid, col_total, col_link = find_header_row_and_cols(ws)
    if header_row is None:
        return []

    data = []
    started = False
    for r in range(header_row + 1, ws.max_row + 1):
        vid = ws.cell(r, col_vid).value
        total = ws.cell(r, col_total).value

        if vid is None and total is None:
            if started:
                break
            else:
                continue

        started = True
        vid_s = str(vid).strip() if vid is not None else ""
        if not vehicle_pattern.match(vid_s):
            continue

        try:
            total_f = float(total) if total is not None else 0.0
        except Exception:
            total_f = 0.0

        data.append((vid_s, total_f, brand))
    return data

# =========================
# Link updater (yours)
# =========================
def relink_vehicle_list_to_global_details(ws):
    header_row, col_vid, _, col_link = find_header_row_and_cols(ws)
    if header_row is None or col_link is None:
        return

    started = False
    for r in range(header_row + 1, ws.max_row + 1):
        vid = ws.cell(r, col_vid).value
        link_cell = ws.cell(r, col_link)

        if vid is None and (link_cell.value is None):
            if started:
                break
            else:
                continue

        started = True
        vid_s = str(vid).strip() if vid is not None else ""
        if not vehicle_pattern.match(vid_s):
            continue

        target = f"V_{vid_s}"
        link_cell.value = link_cell.value or "Open"
        link_cell.hyperlink = f"#'{target}'!A1"
        link_cell.style = "Hyperlink"
        link_cell.alignment = ALIGN_LEFT

# ==========================================================
# Extract Decompte totals (yours)
# ==========================================================
def _norm_txt(x):
    s = "" if x is None else str(x)
    s = s.strip().lower().replace("\u00a0", " ").replace("œ", "oe")
    s = (s.replace("é", "e").replace("è", "e").replace("ê", "e")
         .replace("à", "a").replace("ù", "u"))
    s = re.sub(r"\s+", " ", s)
    return s

def extract_decompte_totals(ws):
    header_row = None
    header_map = {}
    for r in range(1, min(ws.max_row, 40) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        norm = [_norm_txt(v) for v in row_vals]
        if "decompte" in norm:
            header_row = r
            for c, v in enumerate(norm, start=1):
                if v in (
                    "decompte",
                    "total main d'oeuvre", "total main d oeuvre",
                    "total pieces", "total piece",
                    "fgb",
                    "total htv/htva", "total htva", "total htv",
                ):
                    header_map[v] = c
            break

    if header_row:
        for r in range(header_row + 1, ws.max_row + 1):
            d = _norm_txt(ws.cell(r, header_map.get("decompte", 1)).value)
            if d == "total global":
                def get_num(col):
                    if not col:
                        return 0.0
                    v = ws.cell(r, col).value
                    try:
                        return float(v)
                    except Exception:
                        return 0.0

                total_main = get_num(header_map.get("total main d'oeuvre") or header_map.get("total main d oeuvre"))
                total_pieces = get_num(header_map.get("total pieces") or header_map.get("total piece"))
                fgb = get_num(header_map.get("fgb"))
                total_htva = get_num(header_map.get("total htv/htva") or header_map.get("total htva") or header_map.get("total htv"))

                return {
                    "total_main": total_main,
                    "total_pieces": total_pieces,
                    "fgb": fgb,
                    "total_htva": total_htva,
                }

    total_main = None
    total_pieces = None
    fgb = None
    total_htva = None

    for r in range(1, ws.max_row + 1):
        a = _norm_txt(ws.cell(r, 1).value)
        if a in ("total main d'oeuvre", "total main d oeuvre"):
            for c in range(2, min(ws.max_column, 12) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    total_main = float(v)
                    break
        elif a in ("pieces de rechange", "piece de rechange"):
            for c in range(2, min(ws.max_column, 12) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    total_pieces = float(v)
                    break
        elif a == "fgb":
            for c in range(2, min(ws.max_column, 12) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    fgb = float(v)
                    break
        elif a in ("total htv", "total htva", "total htv/htva"):
            for c in range(2, min(ws.max_column, 12) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    total_htva = float(v)
                    break

    return {
        "total_main": 0.0 if total_main is None else total_main,
        "total_pieces": 0.0 if total_pieces is None else total_pieces,
        "fgb": 0.0 if fgb is None else fgb,
        "total_htva": 0.0 if total_htva is None else total_htva,
    }

def create_global_decompte_summary_sheet(out_wb, decompte_refs):
    name = "Global Decompte Summary"
    if name in out_wb.sheetnames:
        i = 1
        while f"{name}_{i}" in out_wb.sheetnames:
            i += 1
        name = f"{name}_{i}"

    ws = out_wb.create_sheet(name)

    ws["A1"] = "Back To Global List"
    ws["A1"].hyperlink = "#'Global Vehicle List'!A1"
    ws["A1"].style = "Hyperlink"
    ws["A1"].alignment = ALIGN_LEFT

    headers = ["Brand", "Source Decompte Sheet", "Total main d'œuvre", "Total pièces", "FGB", "Total HTVA"]
    ws.append([])          # row 2
    ws.append(headers)     # row 3

    for c in range(1, len(headers) + 1):
        cell = ws.cell(3, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border2
        cell.alignment = center

    grand = {"main": 0.0, "pieces": 0.0, "fgb": 0.0, "htva": 0.0}

    for ref in decompte_refs:
        brand = ref["brand"]
        sheet_name = ref["sheet"]

        if sheet_name not in out_wb.sheetnames:
            continue

        src_ws = out_wb[sheet_name]
        totals = extract_decompte_totals(src_ws)

        grand["main"] += totals["total_main"]
        grand["pieces"] += totals["total_pieces"]
        grand["fgb"] += totals["fgb"]
        grand["htva"] += totals["total_htva"]

        r = ws.max_row + 1
        ws.cell(r, 1, brand).border = border2
        ws.cell(r, 2, sheet_name).border = border2
        ws.cell(r, 3, round(totals["total_main"], 3)).border = border2
        ws.cell(r, 4, round(totals["total_pieces"], 3)).border = border2
        ws.cell(r, 5, round(totals["fgb"], 3)).border = border2
        ws.cell(r, 6, round(totals["total_htva"], 3)).border = border2

        ws.cell(r, 2).hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(r, 2).style = "Hyperlink"

        for c in range(1, 7):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)

    tr = ws.max_row + 1
    total_fill = PatternFill("solid", fgColor="C6E0B4")

    ws.cell(tr, 1, "GRAND TOTAL").font = Font(bold=True)
    ws.cell(tr, 1).fill = total_fill
    ws.cell(tr, 1).border = border2

    ws.cell(tr, 2, "").fill = total_fill
    ws.cell(tr, 2).border = border2

    ws.cell(tr, 3, round(grand["main"], 3)).font = Font(bold=True)
    ws.cell(tr, 3).fill = total_fill
    ws.cell(tr, 3).border = border2

    ws.cell(tr, 4, round(grand["pieces"], 3)).font = Font(bold=True)
    ws.cell(tr, 4).fill = total_fill
    ws.cell(tr, 4).border = border2

    ws.cell(tr, 5, round(grand["fgb"], 3)).font = Font(bold=True)
    ws.cell(tr, 5).fill = total_fill
    ws.cell(tr, 5).border = border2

    ws.cell(tr, 6, round(grand["htva"], 3)).font = Font(bold=True)
    ws.cell(tr, 6).fill = total_fill
    ws.cell(tr, 6).border = border2

    for c in range(1, 7):
        ws.cell(tr, c).alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A4"
    set_col_widths(ws, [12, 30, 18, 14, 10, 14])
    return name

# ==========================================================
# V_<VID> builders (your added block)
# ==========================================================
def _norm_label(x):
    s = "" if x is None else str(x)
    s = s.strip().lower().replace("\u00a0", " ")
    s = s.replace("œ", "oe")
    s = s.replace("’", "'")
    s = re.sub(r"\s+", " ", s)
    return s

def extract_tables(ws):
    rows = [[cell.value for cell in r] for r in ws.iter_rows()]

    def find_contains(*needles):
        needles = [_norm_label(n) for n in needles]
        for i, r in enumerate(rows):
            a = _norm_label(r[0] if len(r) else None)
            if any(n in a for n in needles):
                return i
        return None

    i_main = find_contains("main d", "main d'oeuvre", "main d’œuvre", "main d'œuvre")
    i_piece = find_contains("pièce", "piece")

    end_total = len(rows)
    for i, r in enumerate(rows):
        a = _norm_label(r[0] if len(r) else None)
        if "total htva" in a or "total htv" in a:
            end_total = i
            break

    def extract(start, end):
        if start is None:
            return [], []
        header_row = start + 1
        headers = rows[header_row][:8]
        data = []
        for rr in rows[header_row + 1:end]:
            if any(v is not None for v in rr[:8]):
                data.append(rr[:8])
        return headers, data

    main_headers, main_data = extract(i_main, i_piece if i_piece is not None else end_total)
    piece_headers, piece_data = extract(i_piece, end_total)

    return (main_headers, main_data), (piece_headers, piece_data)

def write_table_block(ws, start_row, title, headers, data):
    r = start_row
    ws.cell(r, 1, title).font = label_font
    ws.cell(r, 1).alignment = wrap_left
    r += 1

    if not headers:
        ws.cell(r, 1, "(table not found)").alignment = wrap_left
        return r + 2

    write_row(
        ws, r, headers, start_col=1,
        font=subheader_font, fill=subheader_fill,
        alignment=center, apply_border=True
    )
    r += 1

    for rowvals in data:
        write_row(ws, r, rowvals, start_col=1, alignment=top_left, apply_border=True)
        r += 1

    return r + 2

def paste_sheet_values_only(ws_src, ws_dst, start_row):
    for merged in ws_src.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        ws_dst.merge_cells(
            start_row=min_row + start_row - 1,
            start_column=min_col,
            end_row=max_row + start_row - 1,
            end_column=max_col,
        )

    for row in ws_src.iter_rows():
        for cell in row:
            dst = ws_dst.cell(
                row=cell.row + start_row - 1,
                column=cell.column,
                value=cell.value
            )
            if cell.has_style:
                dst.font = copy(cell.font)
                dst.fill = copy(cell.fill)
                dst.alignment = copy(cell.alignment)
                dst.border = copy(cell.border)
            dst.number_format = cell.number_format
            if cell.hyperlink:
                dst.hyperlink = copy(cell.hyperlink)

    return start_row + ws_src.max_row + 2

# ==========================================================
# ✅ NEW: Designation counts from extracted tables
# ==========================================================
def designation_counts_from_table(headers, data):
    """
    headers: list
    data: list of rows
    Returns: (non_empty_rows, unique_values)
    """
    if not headers or not data:
        return (0, 0)

    def norm(x):
        return ("" if x is None else str(x)).strip().lower().replace("\u00a0", " ")

    # find designation col index
    des_idx = None
    for i, h in enumerate(headers):
        nh = norm(h)
        if nh in ("designation", "désignation") or "designation" in nh or "désignation" in nh:
            des_idx = i
            break

    if des_idx is None:
        return (0, 0)

    vals = []
    for row in data:
        if des_idx >= len(row):
            continue
        v = row[des_idx]
        if v is None:
            continue
        s = str(v).strip()
        if s == "" or s.lower() in ("nan", "none", "-"):
            continue
        vals.append(s)

    return (len(vals), len(set(vals)))

# =========================
# CORE MERGE ENGINE (your build_one_dataset, but parameterized)
# =========================
def build_one_dataset_from_workbooks(brand_to_wb: dict) -> Workbook:
    """
    brand_to_wb: {"TAS": openpyxl.Workbook, "Peugeot": Workbook, "Citroen": Workbook}
    Returns: merged openpyxl Workbook
    """
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    # 1) Global Vehicle List first
    main_ws = out_wb.create_sheet("Global Vehicle List")

    # 2) Copy priority sheets in order
    for priority_name in priority_sheets:
        for brand, wb in brand_to_wb.items():
            if priority_name in wb.sheetnames:
                copy_sheet(wb[priority_name], out_wb, priority_name)
                break

    # 3) Copy all remaining sheets + collect Decompte sheets + map vehicle detail sheets
    brand_sheet_map = {}  # (vid, brand) -> copied_sheet_name
    decompte_refs = []

    for brand, source_wb in brand_to_wb.items():
        for sheet_name in source_wb.sheetnames:
            if sheet_name in priority_sheets:
                continue

            new_name = sheet_name
            counter = 1
            while new_name in out_wb.sheetnames:
                new_name = f"{sheet_name}_{counter}"
                counter += 1

            new_ws = copy_sheet(source_wb[sheet_name], out_wb, new_name)

            base = sheet_name.split("_")[0]
            if vehicle_pattern.match(base):
                brand_sheet_map[(base, brand)] = new_ws.title

            if "decompte" in (sheet_name or "").lower():
                decompte_refs.append({
                    "brand": brand,
                    "sheet": new_ws.title,
                    "source_file": brand
                })

    # 4) Build global list data from copied list sheets
    rows = []
    for brand, sheet_name in vehicle_list_sheets.items():
        if sheet_name not in out_wb.sheetnames:
            continue
        ws = out_wb[sheet_name]
        rows.extend(read_vehicle_list(ws, brand))

    # Aggregate
    agg = {}
    for vid, total, brand in rows:
        if vid not in agg:
            agg[vid] = {"total": 0.0, "brands": set()}
        agg[vid]["total"] += float(total)
        agg[vid]["brands"].add(brand)

    # 5) Write Global Vehicle List
    headers = ["Véhicule ID", "Total HTVA", "Go to sheet", "Brand(s)"]
    main_ws.append(headers)

    for col, h in enumerate(headers, 1):
        c = main_ws.cell(1, col, h)
        c.font = header_font
        c.fill = header_fill
        c.border = border2
        c.alignment = center

    for vid in sorted(agg.keys()):
        total = agg[vid]["total"]
        brands = sorted(agg[vid]["brands"])
        brands_str = ", ".join(brands)
        detail_sheet = f"V_{vid}"

        r = main_ws.max_row + 1
        main_ws.cell(r, 1, vid).border = border2
        main_ws.cell(r, 2, round(total, 3)).border = border2

        link_cell = main_ws.cell(r, 3, "Go to Sheet")
        link_cell.hyperlink = f"#'{detail_sheet}'!A1"
        link_cell.style = "Hyperlink"
        link_cell.border = border2

        main_ws.cell(r, 4, brands_str).border = border2

        for cidx in range(1, 5):
            main_ws.cell(r, cidx).alignment = Alignment(vertical="top", wrap_text=True)

    main_ws.freeze_panes = "A2"
    set_col_widths(main_ws, [14, 14, 14, 24])

    # TOTAL of VEHICLES row
    grand_total = sum(v["total"] for v in agg.values())
    total_row = main_ws.max_row + 2
    total_fill = PatternFill("solid", fgColor="C6E0B4")

    # label
    main_ws.cell(total_row, 1, "TOTAL of VEHICLES").font = Font(bold=True)
    main_ws.cell(total_row, 1).border = border2
    main_ws.cell(total_row, 1).fill = total_fill
    # value
    main_ws.cell(total_row, 2, round(grand_total, 3)).font = Font(bold=True)
    main_ws.cell(total_row, 2).border = border2
    main_ws.cell(total_row, 2).fill = total_fill

    for col in range(1, 5):
        main_ws.cell(total_row, col).alignment = Alignment(vertical="center")

    # 6) Create V_<vid> detail sheets
    for vid in sorted(agg.keys()):
        total = agg[vid]["total"]
        brands = sorted(agg[vid]["brands"])

        ws = out_wb.create_sheet(f"V_{vid}")

        ws["A1"] = "Back To Global List"
        ws["A1"].hyperlink = "#'Global Vehicle List'!A1"
        ws["A1"].style = "Hyperlink"

        ws["A3"] = f"Total HTVA: {round(total, 3)}"
        ws["A3"].font = Font(bold=True, size=12)

        # ✅ NEW: counters across all brands for this VID
        total_main_rows = 0
        total_main_unique = 0
        total_piece_rows = 0
        total_piece_unique = 0

        current_row = 5
        copied_any = False

        for brand in brands:
            source_sheet_name = brand_sheet_map.get((vid, brand))
            if not source_sheet_name:
                continue

            src_ws = out_wb[source_sheet_name]
            (mh, md), (ph, pd_) = extract_tables(src_ws)

            # ✅ NEW: count designations in extracted tables
            mr, mu = designation_counts_from_table(mh, md)
            pr, pu = designation_counts_from_table(ph, pd_)
            total_main_rows += mr
            total_main_unique += mu
            total_piece_rows += pr
            total_piece_unique += pu

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            ws.cell(current_row, 1, brand).font = section_font
            ws.cell(current_row, 1).fill = section_fill
            ws.cell(current_row, 1).alignment = wrap_left
            current_row += 2

            if (mh and md) or (ph and pd_):
                if mh:
                    current_row = write_table_block(ws, current_row, "Main d'œuvre", mh, md)
                if ph:
                    current_row = write_table_block(ws, current_row, "Pièces", ph, pd_)
                copied_any = True
            else:
                ws.cell(current_row, 1, "(Extraction failed → full sheet copied below)").font = Font(italic=True)
                current_row += 1
                current_row = paste_sheet_values_only(src_ws, ws, current_row)
                copied_any = True

        # ✅ NEW: show designation counts summary line
        ws["A4"] = (
            f"Designation rows (Main): {total_main_rows} | Unique: {total_main_unique}    "
            f"Designation rows (Pièce): {total_piece_rows} | Unique: {total_piece_unique}"
        )
        ws["A4"].font = Font(bold=True)

        if not copied_any:
            ws.cell(5, 1, "No detail data found for this vehicle.").font = Font(bold=True, color="FF0000")

        set_col_widths(ws, [10, 14, 14, 8, 46, 6, 12, 12])
        ws.freeze_panes = "A5"

    # 7) Relink the 3 vehicle lists
    for sheet_name in priority_sheets:
        if sheet_name in out_wb.sheetnames:
            relink_vehicle_list_to_global_details(out_wb[sheet_name])
            style_vehicle_list(out_wb[sheet_name])

    # 8) Create Decompte Summary + add link row in Global list
    if decompte_refs:
        decompte_sheet_name = create_global_decompte_summary_sheet(out_wb, decompte_refs)

        link_row = main_ws.max_row + 2

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row_font = Font(bold=True, color="000000")

        for col in range(1, 5):
            c = main_ws.cell(row=link_row, column=col)
            c.border = border2
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.fill = green_fill
            c.font = row_font

        main_ws.cell(row=link_row, column=1, value="Decompte Overall Count")
        main_ws.cell(row=link_row, column=2, value="")

        link_cell = main_ws.cell(row=link_row, column=3, value="Open")
        link_cell.hyperlink = f"#'{decompte_sheet_name}'!A1"
        link_cell.font = row_font
        link_cell.fill = green_fill
        link_cell.border = border2
        link_cell.alignment = Alignment(vertical="top", wrap_text=True)

        main_ws.cell(row=link_row, column=4, value="")

    return out_wb

def build_one_dataset_from_bytes(brand_to_bytes: dict) -> bytes:
    """
    brand_to_bytes: {"TAS": b"...xlsx...", "Peugeot": b"...", "Citroen": b"..."}
    Returns merged workbook bytes.
    """
    brand_to_wb = {}
    for brand, b in brand_to_bytes.items():
        brand_to_wb[brand] = load_workbook(BytesIO(b), data_only=False)

    out_wb = build_one_dataset_from_workbooks(brand_to_wb)

    bio = BytesIO()
    out_wb.save(bio)
    return bio.getvalue()

# Optional legacy helper if you ever want to keep using file paths:
def build_one_dataset_from_paths(file_paths: list[str]) -> bytes:
    brand_to_wb = {}
    for p in file_paths:
        wb = load_workbook(p, data_only=False)
        brand = infer_brand_from_name(p)
        brand_to_wb[brand] = wb

    out_wb = build_one_dataset_from_workbooks(brand_to_wb)
    bio = BytesIO()
    out_wb.save(bio)
    return bio.getvalue()

# ==========================================================
# COMPATIBILITY WRAPPER (fix for Pages/2_Results.py import)
# ==========================================================
def build_global_from_cleaned_bytes(brand_to_bytes: dict) -> bytes:
    """
    Compatibility wrapper expected by Pages/2_Results.py
    """
    return build_one_dataset_from_bytes(brand_to_bytes)