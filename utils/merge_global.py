# utils/merge_global.py
# -*- coding: utf-8 -*-

from __future__ import annotations

import re
from io import BytesIO
from copy import copy
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# CONFIG
# =========================================================
priority_sheets = [
    "TAS Vehicle List",
    "Peugeot Vehicle List",
    "Citroen Vehicle List",
]

vehicle_list_sheets = {
    "TAS": "TAS Vehicle List",
    "Peugeot": "Peugeot Vehicle List",
    "Citroen": "Citroen Vehicle List",
}

vehicle_pattern = re.compile(r"^\d{2}-\d{6}$")


def infer_brand_from_name(name: str) -> str:
    p = (name or "").lower()
    if "tas" in p:
        return "TAS"
    if "peugeot" in p:
        return "Peugeot"
    if "citroen" in p or "citreon" in p:
        return "Citroen"
    return "Unknown"


# =========================================================
# STYLE CONSTANTS (clean consistent theme)
# =========================================================
THIN = Side(style="thin", color="1F1F1F")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Theme colors
COLOR_HEADER = "1F4E79"   # dark blue
COLOR_SECTION = "2F5597"  # medium blue
COLOR_TOTAL = "C6E0B4"    # green total
COLOR_MAIN_HDR = "F4C7A1" # light orange
COLOR_PIECE_HDR = "CFE2F3"# light blue

FILL_LIST_HEADER = PatternFill("solid", fgColor=COLOR_HEADER)
FILL_GLOBAL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER)
FILL_SECTION = PatternFill("solid", fgColor=COLOR_SECTION)
FILL_TOTAL = PatternFill("solid", fgColor=COLOR_TOTAL)
FILL_MAIN_HEADER = PatternFill("solid", fgColor=COLOR_MAIN_HDR)
FILL_PIECE_HEADER = PatternFill("solid", fgColor=COLOR_PIECE_HDR)

FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
FONT_BOLD = Font(bold=True)
FONT_SECTION = Font(bold=True, color="FFFFFF", size=12)
FONT_TITLE = Font(bold=True, color="FFFFFF", size=13)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _normalize(s) -> str:
    return str(s).strip().lower().replace("\u00a0", " ") if s is not None else ""


def _apply_border_range(ws, min_row, max_row, min_col, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = BORDER_THIN


def set_col_widths(ws, widths: List[int]):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# =========================================================
# Copy sheet EXACTLY (values + styles + hyperlinks)
# =========================================================
def copy_sheet(source_sheet, master_wb, new_title):
    new_sheet = master_wb.create_sheet(title=new_title)

    for col, dim in source_sheet.column_dimensions.items():
        new_sheet.column_dimensions[col] = copy(dim)

    for row, dim in source_sheet.row_dimensions.items():
        new_sheet.row_dimensions[row] = copy(dim)

    for merged in source_sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged))

    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.alignment = copy(cell.alignment)
                new_cell.border = copy(cell.border)
            if cell.hyperlink:
                new_cell.hyperlink = copy(cell.hyperlink)
            if cell.number_format:
                new_cell.number_format = cell.number_format

    if source_sheet.freeze_panes:
        new_sheet.freeze_panes = source_sheet.freeze_panes

    new_sheet.sheet_properties.tabColor = source_sheet.sheet_properties.tabColor
    return new_sheet


# =========================================================
# Vehicle list header detection
# =========================================================
def find_header_row_and_cols(ws):
    max_row = min(ws.max_row, 50)
    max_col = min(ws.max_column, 30)

    for r in range(1, max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        norm = [_normalize(v) for v in vals]

        if any("vehicule id" in v or "vÃ©hicule id" in v for v in norm):
            col_vid = None
            col_total = None
            col_link = None

            for c, v in enumerate(norm, start=1):
                if col_vid is None and ("vehicule id" in v or "vÃ©hicule id" in v):
                    col_vid = c
                if col_total is None and ("total" in v and ("htva" in v or "htv" in v or v.strip() == "total")):
                    col_total = c
                if col_link is None and ("open" == v.strip() or "go to sheet" in v or "open sheet" in v):
                    col_link = c

            if col_vid is None:
                col_vid = 1
            if col_total is None:
                col_total = 2
            if col_link is None:
                col_link = 3 if ws.max_column >= 3 else None

            return r, col_vid, col_total, col_link

    return None, None, None, None


def _to_float_fr(v) -> float:
    # robust: handles 7 437,367 / "7437,367" / numeric
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return 0.0

    s = str(v).strip()
    if not s:
        return 0.0
    if s.startswith("="):  # formula text (should use data_only workbook)
        return 0.0

    s = s.replace("\u00A0", " ").replace("\u202F", " ").replace(" ", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def read_vehicle_list(ws, brand: str):
    header_row, col_vid, col_total, col_link = find_header_row_and_cols(ws)
    if header_row is None:
        return []

    data = []
    started = False

    for r in range(header_row + 1, ws.max_row + 1):
        vid = ws.cell(r, col_vid).value
        total = ws.cell(r, col_total).value

        if vid is None and total is None:
            if started:
                break
            continue
        started = True

        vid_s = str(vid).strip() if vid is not None else ""
        if not vehicle_pattern.match(vid_s):
            continue

        total_f = _to_float_fr(total)
        data.append((vid_s, total_f, brand))

    return data


# =========================================================
# Style helpers for global list + sections
# =========================================================
def style_global_vehicle_list(ws):
    # row 1 header
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        cell.fill = FILL_GLOBAL_HEADER
        cell.font = FONT_WHITE_BOLD
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    # borders for table
    _apply_border_range(ws, 1, ws.max_row, 1, max_col)

    # alignments
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).alignment = ALIGN_LEFT
        ws.cell(r, 2).alignment = ALIGN_RIGHT
        ws.cell(r, 3).alignment = ALIGN_LEFT
        ws.cell(r, 4).alignment = ALIGN_LEFT


def add_total_of_vehicles_row(main_ws, total_value: float, ncols: int = 4):
    r = main_ws.max_row + 2  # blank line then total row

    # A label
    cA = main_ws.cell(r, 1, "TOTAL of VEHICLES")
    cA.font = FONT_WHITE_BOLD
    cA.fill = FILL_SECTION
    cA.border = BORDER_THIN
    cA.alignment = ALIGN_LEFT

    # B value
    cB = main_ws.cell(r, 2, round(float(total_value), 3))
    cB.font = Font(bold=True, color="000000")
    cB.fill = FILL_TOTAL
    cB.border = BORDER_THIN
    cB.alignment = ALIGN_CENTER

    # fill remaining cells in green
    for col in range(3, ncols + 1):
        c = main_ws.cell(r, col, "")
        c.fill = FILL_TOTAL
        c.border = BORDER_THIN
        c.alignment = ALIGN_CENTER

    return r


# =========================================================
# Decompte summary (kept from your logic, stable)
# =========================================================
def _norm_txt(x):
    s = "" if x is None else str(x)
    s = s.strip().lower().replace("\u00a0", " ").replace("Å“", "oe")
    s = (
        s.replace("Ã©", "e")
        .replace("Ã¨", "e")
        .replace("Ãª", "e")
        .replace("Ã ", "a")
        .replace("Ã¹", "u")
    )
    s = re.sub(r"\s+", " ", s)
    return s


def extract_decompte_totals(ws):
    header_row = None
    header_map = {}

    for r in range(1, min(ws.max_row, 40) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        norm = [_norm_txt(v) for v in row_vals]
        if "decompte" in norm:
            header_row = r
            for c, v in enumerate(norm, start=1):
                if v in (
                    "decompte",
                    "total main d'oeuvre", "total main d oeuvre",
                    "total pieces", "total piece",
                    "fgb",
                    "total htv/htva", "total htva", "total htv",
                ):
                    header_map[v] = c
            break

    if header_row:
        for r in range(header_row + 1, ws.max_row + 1):
            d = _norm_txt(ws.cell(r, header_map.get("decompte", 1)).value)
            if d == "total global":
                def get_num(col):
                    if not col:
                        return 0.0
                    return _to_float_fr(ws.cell(r, col).value)

                total_main = get_num(header_map.get("total main d'oeuvre") or header_map.get("total main d oeuvre"))
                total_pieces = get_num(header_map.get("total pieces") or header_map.get("total piece"))
                fgb = get_num(header_map.get("fgb"))
                total_htva = get_num(header_map.get("total htv/htva") or header_map.get("total htva") or header_map.get("total htv"))

                return {"total_main": total_main, "total_pieces": total_pieces, "fgb": fgb, "total_htva": total_htva}

    return {"total_main": 0.0, "total_pieces": 0.0, "fgb": 0.0, "total_htva": 0.0}


def create_global_decompte_summary_sheet(out_wb, decompte_refs):
    name = "Global Decompte Summary"
    if name in out_wb.sheetnames:
        i = 1
        while f"{name}_{i}" in out_wb.sheetnames:
            i += 1
        name = f"{name}_{i}"

    ws = out_wb.create_sheet(name)

    ws["A1"] = "Back To Global List"
    ws["A1"].hyperlink = "#'Global Vehicle List'!A1"
    ws["A1"].style = "Hyperlink"
    ws["A1"].alignment = ALIGN_LEFT

    headers = ["Brand", "Source Decompte Sheet", "Total main d'oeuvre", "Total pieces", "FGB", "Total HTVA"]
    ws.append([])      # row 2
    ws.append(headers) # row 3

    for c in range(1, len(headers) + 1):
        cell = ws.cell(3, c)
        cell.font = FONT_WHITE_BOLD
        cell.fill = FILL_GLOBAL_HEADER
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER

    grand = {"main": 0.0, "pieces": 0.0, "fgb": 0.0, "htva": 0.0}

    for ref in decompte_refs:
        brand = ref["brand"]
        sheet_name = ref["sheet"]
        if sheet_name not in out_wb.sheetnames:
            continue

        src_ws = out_wb[sheet_name]
        totals = extract_decompte_totals(src_ws)

        grand["main"] += totals["total_main"]
        grand["pieces"] += totals["total_pieces"]
        grand["fgb"] += totals["fgb"]
        grand["htva"] += totals["total_htva"]

        r = ws.max_row + 1
        ws.cell(r, 1, brand).border = BORDER_THIN
        ws.cell(r, 2, sheet_name).border = BORDER_THIN
        ws.cell(r, 3, round(totals["total_main"], 3)).border = BORDER_THIN
        ws.cell(r, 4, round(totals["total_pieces"], 3)).border = BORDER_THIN
        ws.cell(r, 5, round(totals["fgb"], 3)).border = BORDER_THIN
        ws.cell(r, 6, round(totals["total_htva"], 3)).border = BORDER_THIN

        ws.cell(r, 2).hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(r, 2).style = "Hyperlink"

        for c in range(1, 7):
            ws.cell(r, c).alignment = ALIGN_TOP_LEFT

    tr = ws.max_row + 1
    ws.cell(tr, 1, "GRAND TOTAL").font = FONT_BOLD
    ws.cell(tr, 1).fill = FILL_TOTAL
    ws.cell(tr, 1).border = BORDER_THIN

    ws.cell(tr, 2, "").fill = FILL_TOTAL
    ws.cell(tr, 2).border = BORDER_THIN

    ws.cell(tr, 3, round(grand["main"], 3)).font = FONT_BOLD
    ws.cell(tr, 3).fill = FILL_TOTAL
    ws.cell(tr, 3).border = BORDER_THIN

    ws.cell(tr, 4, round(grand["pieces"], 3)).font = FONT_BOLD
    ws.cell(tr, 4).fill = FILL_TOTAL
    ws.cell(tr, 4).border = BORDER_THIN

    ws.cell(tr, 5, round(grand["fgb"], 3)).font = FONT_BOLD
    ws.cell(tr, 5).fill = FILL_TOTAL
    ws.cell(tr, 5).border = BORDER_THIN

    ws.cell(tr, 6, round(grand["htva"], 3)).font = FONT_BOLD
    ws.cell(tr, 6).fill = FILL_TOTAL
    ws.cell(tr, 6).border = BORDER_THIN

    ws.freeze_panes = "A4"
    set_col_widths(ws, [14, 30, 20, 16, 12, 16])
    _apply_border_range(ws, 3, ws.max_row, 1, 6)
    return name


# =========================================================
# ✅ V_<VID> detail sheet builder (organized layout)
# =========================================================
def _read_brand_detail_sheet_values(ws_src) -> Tuple[List[str], List[List]]:
    """
    Extracts blocks from a brand per-vehicle sheet:
    - finds "Main d'oeuvre" section
    - finds "Pièce" section
    Returns: (main_block_rows, piece_block_rows) as raw rows (values).
    If extraction fails, returns empty lists.
    """
    rows = [[cell.value for cell in r] for r in ws_src.iter_rows()]

    def find_row_index_contains(*needles):
        needles = [str(n).lower() for n in needles]
        for i, r in enumerate(rows):
            a = "" if not r else ("" if r[0] is None else str(r[0]).lower())
            if any(n in a for n in needles):
                return i
        return None

    i_main = find_row_index_contains("main d", "main d'oeuvre", "main d'")  # tolerant
    i_piece = find_row_index_contains("pièce", "piece")

    # Find total row as an end marker
    i_total = None
    for i, r in enumerate(rows):
        a = "" if not r else ("" if r[0] is None else str(r[0]).lower())
        if "total htva" in a or "total htv" in a:
            i_total = i
            break

    if i_main is None:
        return [], []
    if i_total is None:
        i_total = len(rows)

    if i_piece is None:
        i_piece = i_total

    def slice_block(start_title_idx, end_idx) -> List[List]:
        # expect: title row, header row, data...
        header_idx = start_title_idx + 1
        if header_idx >= len(rows):
            return []
        out = []
        for rr in rows[header_idx: end_idx]:
            if any(v is not None and str(v).strip() != "" for v in rr[:12]):
                out.append(rr[:12])
        return out

    main_block = slice_block(i_main, i_piece)
    piece_block = slice_block(i_piece, i_total) if i_piece < i_total else []
    return main_block, piece_block


def _count_designations(block_rows: List[List]) -> Tuple[int, int]:
    """
    Count non-empty designation rows + unique values in a block.
    We try to find a column whose header contains "designation".
    """
    if not block_rows or len(block_rows) < 2:
        return (0, 0)

    header = block_rows[0]
    data = block_rows[1:]

    def norm(x):
        return ("" if x is None else str(x)).strip().lower()

    des_idx = None
    for i, h in enumerate(header):
        if "designation" in norm(h) or "désignation" in norm(h):
            des_idx = i
            break

    if des_idx is None:
        return (0, 0)

    vals = []
    for r in data:
        if des_idx >= len(r):
            continue
        v = r[des_idx]
        s = "" if v is None else str(v).strip()
        if not s or s.lower() in {"nan", "none", "-"}:
            continue
        vals.append(s)

    return (len(vals), len(set(vals)))


def _write_block(ws, start_row: int, title: str, header: List, data: List[List], header_fill: PatternFill) -> int:
    # Title
    ws.cell(start_row, 1, title).font = Font(bold=True, size=14)
    ws.cell(start_row, 1).alignment = ALIGN_LEFT
    start_row += 1

    if not header:
        ws.cell(start_row, 1, "(no data)").font = Font(italic=True, color="666666")
        return start_row + 2

    # Header
    for c, h in enumerate(header, start=1):
        cell = ws.cell(start_row, c, h)
        cell.font = FONT_BOLD
        cell.fill = header_fill
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    start_row += 1

    # Data
    for row_vals in data:
        for c, v in enumerate(row_vals, start=1):
            cell = ws.cell(start_row, c, v)
            cell.alignment = ALIGN_TOP_LEFT if not isinstance(v, (int, float)) else ALIGN_RIGHT
            cell.border = BORDER_THIN
        start_row += 1

    return start_row + 2


def _style_right_panel(ws, col_label: int, col_val: int, top_row: int):
    # Panel title
    ws.merge_cells(start_row=top_row, start_column=col_label, end_row=top_row, end_column=col_val)
    t = ws.cell(top_row, col_label, "Designation Stats")
    t.font = FONT_WHITE_BOLD
    t.fill = FILL_SECTION
    t.alignment = ALIGN_CENTER
    t.border = BORDER_THIN

    # widths
    ws.column_dimensions[get_column_letter(col_label)].width = 22
    ws.column_dimensions[get_column_letter(col_val)].width = 10

    for r in range(top_row, top_row + 6):
        for c in range(col_label, col_val + 1):
            ws.cell(r, c).border = BORDER_THIN


def _build_vehicle_sheet(out_wb: Workbook, vid: str, total: float, brands: List[str], brand_sheet_map: Dict[Tuple[str, str], str]):
    sheet_name = f"V_{vid}"
    ws = out_wb.create_sheet(sheet_name)

    # Back link
    ws["A1"] = "Back To Global List"
    ws["A1"].hyperlink = "#'Global Vehicle List'!A1"
    ws["A1"].style = "Hyperlink"
    ws["A1"].alignment = ALIGN_LEFT

    # Header line
    ws["A3"] = f"Total HTVA: {round(float(total), 3)}"
    ws["A3"].font = Font(bold=True, size=12)

    # Fixed right panel columns (K:L)
    panel_label_col = 11  # K
    panel_val_col = 12    # L
    panel_top_row = 3
    _style_right_panel(ws, panel_label_col, panel_val_col, panel_top_row)

    # Counters
    main_rows_total = 0
    main_unique_total = 0
    piece_rows_total = 0
    piece_unique_total = 0

    # Body starts
    r = 5

    for brand in brands:
        src_sheet_name = brand_sheet_map.get((vid, brand))
        if not src_sheet_name or src_sheet_name not in out_wb.sheetnames:
            continue

        src_ws = out_wb[src_sheet_name]

        # Brand section bar
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        bcell = ws.cell(r, 1, brand)
        bcell.font = FONT_SECTION
        bcell.fill = FILL_SECTION
        bcell.alignment = ALIGN_LEFT
        bcell.border = BORDER_THIN
        r += 2

        main_block, piece_block = _read_brand_detail_sheet_values(src_ws)

        # If we successfully extracted structured blocks:
        if main_block:
            header = main_block[0]
            data = main_block[1:]
            mr, mu = _count_designations(main_block)
            main_rows_total += mr
            main_unique_total += mu
            r = _write_block(ws, r, "Main d'oeuvre", header, data, FILL_MAIN_HEADER)

        if piece_block:
            header = piece_block[0]
            data = piece_block[1:]
            pr, pu = _count_designations(piece_block)
            piece_rows_total += pr
            piece_unique_total += pu
            r = _write_block(ws, r, "Pièces", header, data, FILL_PIECE_HEADER)

        if not main_block and not piece_block:
            ws.cell(r, 1, "(Could not parse blocks from source sheet)").font = Font(italic=True, color="666666")
            r += 2

    # Fill right panel values
    ws.cell(panel_top_row + 1, panel_label_col, "Main d'oeuvre").alignment = ALIGN_LEFT
    ws.cell(panel_top_row + 1, panel_label_col).border = BORDER_THIN
    ws.cell(panel_top_row + 1, panel_val_col, int(main_unique_total)).alignment = ALIGN_CENTER
    ws.cell(panel_top_row + 1, panel_val_col).border = BORDER_THIN

    ws.cell(panel_top_row + 2, panel_label_col, "Pièces").alignment = ALIGN_LEFT
    ws.cell(panel_top_row + 2, panel_label_col).border = BORDER_THIN
    ws.cell(panel_top_row + 2, panel_val_col, int(piece_unique_total)).alignment = ALIGN_CENTER
    ws.cell(panel_top_row + 2, panel_val_col).border = BORDER_THIN

    ws.cell(panel_top_row + 4, panel_label_col, "Rows Main").alignment = ALIGN_LEFT
    ws.cell(panel_top_row + 4, panel_label_col).border = BORDER_THIN
    ws.cell(panel_top_row + 4, panel_val_col, int(main_rows_total)).alignment = ALIGN_CENTER
    ws.cell(panel_top_row + 4, panel_val_col).border = BORDER_THIN

    ws.cell(panel_top_row + 5, panel_label_col, "Rows Pièces").alignment = ALIGN_LEFT
    ws.cell(panel_top_row + 5, panel_label_col).border = BORDER_THIN
    ws.cell(panel_top_row + 5, panel_val_col, int(piece_rows_total)).alignment = ALIGN_CENTER
    ws.cell(panel_top_row + 5, panel_val_col).border = BORDER_THIN

    # widths for main table area
    set_col_widths(ws, [12, 14, 14, 10, 46, 10, 14, 14])
    ws.freeze_panes = "A5"


# =========================================================
# ✅ MAIN MERGE: builds Global Vehicle List + ALWAYS creates V_<VID>
# =========================================================
def build_one_dataset_from_workbooks(brand_to_wb_copy: dict, brand_to_wb_values: dict) -> Workbook:
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    # 1) Create global list first
    main_ws = out_wb.create_sheet("Global Vehicle List")

    # 2) Copy priority list sheets
    for priority_name in priority_sheets:
        for brand, wb in brand_to_wb_copy.items():
            if priority_name in wb.sheetnames:
                copy_sheet(wb[priority_name], out_wb, priority_name)
                break

    # 3) Copy all remaining sheets + map brand vehicle details
    brand_sheet_map: Dict[Tuple[str, str], str] = {}
    decompte_refs = []

    for brand, source_wb in brand_to_wb_copy.items():
        for sheet_name in source_wb.sheetnames:
            if sheet_name in priority_sheets:
                continue

            new_name = sheet_name
            counter = 1
            while new_name in out_wb.sheetnames:
                new_name = f"{sheet_name}_{counter}"
                counter += 1

            new_ws = copy_sheet(source_wb[sheet_name], out_wb, new_name)

            base = sheet_name.split("_")[0]
            if vehicle_pattern.match(base):
                brand_sheet_map[(base, brand)] = new_ws.title

            if "decompte" in (sheet_name or "").lower():
                decompte_refs.append({"brand": brand, "sheet": new_ws.title, "source_file": brand})

    # 4) Read totals from VALUES workbooks (data_only=True)
    rows = []
    for brand, sheet_name in vehicle_list_sheets.items():
        if brand not in brand_to_wb_values:
            continue
        wb_vals = brand_to_wb_values[brand]
        if sheet_name not in wb_vals.sheetnames:
            continue
        ws_vals = wb_vals[sheet_name]
        rows.extend(read_vehicle_list(ws_vals, brand))

    # Aggregate per vehicle id across brands
    agg: Dict[str, Dict] = {}
    for vid, total, brand in rows:
        if vid not in agg:
            agg[vid] = {"total": 0.0, "brands": set()}
        agg[vid]["total"] += float(total)
        agg[vid]["brands"].add(brand)

    # 5) Write Global Vehicle List with links
    headers = ["Véhicule ID", "Total HTVA", "Go to sheet", "Brand(s)"]
    main_ws.append(headers)

    # Create detail sheets first (so links are ALWAYS valid)
    for vid in sorted(agg.keys()):
        total = float(agg[vid]["total"])
        brands = sorted(list(agg[vid]["brands"]))
        _build_vehicle_sheet(out_wb, vid, total, brands, brand_sheet_map)

    # Now write rows + hyperlink
    grand_total_vehicles = 0.0
    for vid in sorted(agg.keys()):
        total = float(agg[vid]["total"])
        grand_total_vehicles += total
        brands = sorted(list(agg[vid]["brands"]))
        brands_str = ", ".join(brands)

        r = main_ws.max_row + 1
        main_ws.cell(r, 1, vid).border = BORDER_THIN
        main_ws.cell(r, 2, round(total, 3)).border = BORDER_THIN
        main_ws.cell(r, 4, brands_str).border = BORDER_THIN

        # ✅ Hyperlink fix: target MUST exist.
        # We create V_<VID> above, so this should always work.
        target_v = f"V_{vid}"
        link_cell = main_ws.cell(r, 3, "Go to Sheet")
        if target_v in out_wb.sheetnames:
            link_cell.hyperlink = f"#'{target_v}'!A1"
        elif vid in out_wb.sheetnames:
            # fallback if someone disables V_ sheets
            link_cell.hyperlink = f"#'{vid}'!A1"
        else:
            # last resort: keep text but no hyperlink
            link_cell.value = "Missing sheet"
            link_cell.hyperlink = None

        link_cell.style = "Hyperlink"
        link_cell.border = BORDER_THIN

        for cidx in range(1, 5):
            main_ws.cell(r, cidx).alignment = ALIGN_TOP_LEFT

    main_ws.freeze_panes = "A2"
    set_col_widths(main_ws, [16, 14, 14, 24])

    # Style global list (headers etc.)
    style_global_vehicle_list(main_ws)

    # Total row
    add_total_of_vehicles_row(main_ws, grand_total_vehicles, ncols=4)

    # 6) Decompte summary
    if decompte_refs:
        decompte_sheet_name = create_global_decompte_summary_sheet(out_wb, decompte_refs)

        link_row = main_ws.max_row + 2
        for col in range(1, 5):
            c = main_ws.cell(row=link_row, column=col)
            c.border = BORDER_THIN
            c.alignment = ALIGN_TOP_LEFT
            c.fill = FILL_TOTAL
            c.font = FONT_BOLD

        main_ws.cell(row=link_row, column=1, value="Decompte Overall Count")
        link_cell = main_ws.cell(row=link_row, column=3, value="Open")
        link_cell.hyperlink = f"#'{decompte_sheet_name}'!A1"
        link_cell.style = "Hyperlink"
        link_cell.border = BORDER_THIN
        link_cell.alignment = ALIGN_TOP_LEFT
        link_cell.fill = FILL_TOTAL
        link_cell.font = FONT_BOLD

    return out_wb


def build_one_dataset_from_bytes(brand_to_bytes: dict) -> bytes:
    brand_to_wb_copy = {}
    brand_to_wb_values = {}

    for brand, b in brand_to_bytes.items():
        bio = BytesIO(b)
        brand_to_wb_copy[brand] = load_workbook(bio, data_only=False)
        bio.seek(0)
        brand_to_wb_values[brand] = load_workbook(bio, data_only=True)

    out_wb = build_one_dataset_from_workbooks(brand_to_wb_copy, brand_to_wb_values)

    out = BytesIO()
    out_wb.save(out)
    return out.getvalue()


def build_one_dataset_from_paths(file_paths: list[str]) -> bytes:
    brand_to_wb_copy = {}
    brand_to_wb_values = {}

    for p in file_paths:
        brand = infer_brand_from_name(p)
        brand_to_wb_copy[brand] = load_workbook(p, data_only=False)
        brand_to_wb_values[brand] = load_workbook(p, data_only=True)

    out_wb = build_one_dataset_from_workbooks(brand_to_wb_copy, brand_to_wb_values)
    out = BytesIO()
    out_wb.save(out)
    return out.getvalue()


def build_global_from_cleaned_bytes(brand_to_bytes: dict) -> bytes:
    return build_one_dataset_from_bytes(brand_to_bytes)