# utils/dashboards_core.py
# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
import unicodedata
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.io as pio
import streamlit as st
from openpyxl import load_workbook


# =========================================================
# Designation builder helpers (TOP-LEVEL)
# =========================================================
def _parse_hyperlink_sheet_name(target: str) -> str | None:
    """
    Parses hyperlink target like:
      "#'Designation Count - Main d''oeuvre'!A1"
      "#'Designation Count - Pièce'!A1"
      "#Designation Count - Pièce!A1"
    Returns the sheet name.
    """
    if not target:
        return None
    t = str(target).strip()

    if "!" not in t:
        return None

    left = t.split("!", 1)[0]  # "#'Sheet Name'" or "#Sheet Name"
    left = left.lstrip("#").strip()

    # remove wrapping quotes if present
    if left.startswith("'") and left.endswith("'"):
        left = left[1:-1]

    # excel doubles single quotes inside names
    left = left.replace("''", "'").strip()
    return left or None


def build_df_design_from_dataset(ds_bytes: bytes, pick_col_fn) -> pd.DataFrame:
    """
    Builds df_design with columns:
      Brand | Type | Designation | Count

    Reads the designation sheets referenced by hyperlinks inside each brand Vehicle List sheet:
      - TAS Vehicle List
      - Peugeot Vehicle List
      - Citroen Vehicle List
    """
    if not ds_bytes:
        return pd.DataFrame(columns=["Brand", "Type", "Designation", "Count"])

    wb = load_workbook(BytesIO(ds_bytes), data_only=True)

    brand_list_sheets = [
        ("TAS", "TAS Vehicle List"),
        ("Peugeot", "Peugeot Vehicle List"),
        ("Citroen", "Citroen Vehicle List"),
    ]

    out_rows: list[pd.DataFrame] = []

    def find_linked_designation_sheet(ws, contains_text: str) -> str | None:
        needle = str(contains_text).lower()
        for r in range(1, min(ws.max_row, 3000) + 1):
            a = ws.cell(r, 1).value
            if a is None:
                continue
            if needle in str(a).lower():
                link_cell = ws.cell(r, 3)  # column C has "Go to ..."
                if link_cell.hyperlink and link_cell.hyperlink.target:
                    return _parse_hyperlink_sheet_name(link_cell.hyperlink.target)
        return None

    for brand, list_sheet in brand_list_sheets:
        if list_sheet not in wb.sheetnames:
            continue

        ws = wb[list_sheet]

        main_sheet = find_linked_designation_sheet(ws, "Designation Count - Main")
        piece_sheet = find_linked_designation_sheet(ws, "Designation Count - Pièce")
        if not piece_sheet:
            piece_sheet = find_linked_designation_sheet(ws, "Designation Count - Piece")

        for typ, sh in [("Main d'œuvre", main_sheet), ("Pièce", piece_sheet)]:
            if not sh or sh not in wb.sheetnames:
                continue

            try:
                df = pd.read_excel(BytesIO(ds_bytes), sheet_name=sh, header=2)
            except Exception:
                continue

            if df is None or df.empty:
                continue

            c_des = pick_col_fn(df, ["designation", "désignation"])
            c_cnt = pick_col_fn(df, ["main count", "piece count", "pièce count", "count", "nb", "nombre"])
            if not c_des or not c_cnt:
                continue

            tmp = df[[c_des, c_cnt]].copy()
            tmp.columns = ["Designation", "Count"]
            tmp["Designation"] = tmp["Designation"].astype(str).str.strip()
            tmp["Count"] = pd.to_numeric(tmp["Count"], errors="coerce").fillna(0).astype(int)

            tmp = tmp[tmp["Designation"].ne("")].copy()
            tmp = tmp[~tmp["Designation"].str.contains("total", case=False, na=False)].copy()

            tmp["Brand"] = brand
            tmp["Type"] = typ

            out_rows.append(tmp[["Brand", "Type", "Designation", "Count"]])

    wb.close()

    if not out_rows:
        return pd.DataFrame(columns=["Brand", "Type", "Designation", "Count"])

    df_design = pd.concat(out_rows, ignore_index=True)
    df_design["Type"] = df_design["Type"].replace({"Pièces": "Pièce"}).astype(str).str.strip()
    return df_design


# =========================================================
# Main dashboard
# =========================================================
def render_dashboards() -> None:
    """
    Dashboard tabs:
      1) Overview
      2) Brand Summary (Vehicles / FGB / Décompte)
      3) Age vs Main/Pieces Times Done vs Cost (Designation Stats)
      4) Brand Operations (Pièces & Main d'œuvre)
    """

    # Safe page config
    try:
        st.set_page_config(page_title="Dashboard", layout="wide")
    except Exception:
        pass

    # Plotly base template
    pio.templates.default = "plotly_white"

    # -----------------------------
    # Theme: Pink / Green / Blue
    # -----------------------------
    BLUE = "#2563EB"
    PINK = "#EC4899"
    GREEN = "#16A34A"
    BLUE_L = "#60A5FA"
    PINK_L = "#F472B6"
    GREEN_L = "#22C55E"

    GRID_COL = "rgba(37,99,235,0.10)"
    FONT_COL = "#000000"

    # Plotly default colors (discrete)
    px.defaults.color_discrete_sequence = [
        BLUE,
        PINK,
        GREEN,
        BLUE_L,
        PINK_L,
        GREEN_L,
    ]

    # -----------------------------
    # GLOBAL CSS: BIG HEADINGS + BLACK TEXT
    # -----------------------------
    st.markdown(
        f"""
        <style>
          .stApp {{ background: #f7f7fb; }}
          section.main > div {{ padding-top: 1rem; }}

          h1 {{ font-size: 44px !important; font-weight: 950 !important; color:{FONT_COL} !important; }}
          h2 {{ font-size: 34px !important; font-weight: 950 !important; color:{FONT_COL} !important; }}
          h3 {{ font-size: 26px !important; font-weight: 950 !important; color:{FONT_COL} !important; }}
          h4, h5, h6 {{ color:{FONT_COL} !important; font-weight: 900 !important; }}

          .stMarkdown, .stMarkdown p, .stText, .stCaption, .stAlert,
          label, span, small, div, p, li {{
            color: {FONT_COL} !important;
          }}

          div[data-baseweb="tab-list"] button {{ font-weight: 900 !important; color: {FONT_COL} !important; }}
          div[data-baseweb="tab-list"] {{ gap: 10px; }}

          .card {{
            background: #ffffff;
            border: 1px solid #e9eaf5;
            border-radius: 16px;
            padding: 16px 16px;
            box-shadow: 0 10px 24px rgba(2, 6, 23, 0.06);
          }}

          .card-title {{
            font-size: 17px !important;
            font-weight: 950 !important;
            color: {FONT_COL} !important;
            margin-bottom: 2px;
          }}
          .card-sub {{
            font-size: 13px !important;
            font-weight: 750 !important;
            color: {FONT_COL} !important;
            margin-bottom: 10px;
          }}

          .kpi {{
            font-size: 34px;
            font-weight: 950;
            color: {FONT_COL} !important;
            line-height: 1.1;
            margin: 0;
          }}
          .kpi-big {{
            font-size: 56px;
            font-weight: 950;
            color: {FONT_COL} !important;
            line-height: 1.0;
            margin: 0;
          }}

          .pink   {{ color: {PINK} !important; }}
          .blue   {{ color: {BLUE} !important; }}
          .green  {{ color: {GREEN} !important; }}
          .muted  {{ color: {FONT_COL} !important; }}

          .js-plotly-plot .plotly, .js-plotly-plot .plot-container {{ background: transparent !important; }}
          .stDataFrame, .stDataFrame * {{ color:{FONT_COL} !important; }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    results = st.session_state.get("results", {})
    if not isinstance(results, dict) or not results:
        st.info("No results found. Select a session or run Cleaning first.")
        return

    CURRENCY = "TND"

    # -----------------------------
    # Helpers
    # -----------------------------
    def _wrap_card_start(title: str | None = None, subtitle: str | None = None):
        st.markdown('<div class="card">', unsafe_allow_html=True)
        if title:
            st.markdown(f'<div class="card-title">{title}</div>', unsafe_allow_html=True)
        if subtitle:
            st.markdown(f'<div class="card-sub">{subtitle}</div>', unsafe_allow_html=True)

    def _wrap_card_end():
        st.markdown("</div>", unsafe_allow_html=True)

    def card_kpi(title: str, value: str, subtitle: str = "", color_class: str = "muted", big: bool = False):
        size_class = "kpi-big" if big else "kpi"
        st.markdown(
            f"""
            <div class="card">
              <div class="card-title">{title}</div>
              <div class="card-sub">{subtitle}</div>
              <div class="{size_class} {color_class}">{value}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    def _fig_black_text(fig):
        fig.update_layout(
            font=dict(color=FONT_COL),
            legend=dict(font=dict(color=FONT_COL), title_font=dict(color=FONT_COL)),
        )
        fig.update_xaxes(title_font=dict(color=FONT_COL), tickfont=dict(color=FONT_COL), gridcolor=GRID_COL)
        fig.update_yaxes(title_font=dict(color=FONT_COL), tickfont=dict(color=FONT_COL), gridcolor=GRID_COL)
        return fig

    def _layout_plotly_card(fig, height: int, title_pad: int = 10, showlegend: bool = False):
        fig.update_layout(
            height=height,
            margin=dict(l=10, r=10, t=title_pad, b=10),
            showlegend=showlegend,
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
        )
        _fig_black_text(fig)
        return fig

    def _strip_accents(s: str) -> str:
        s = str(s).replace("œ", "oe").replace("Œ", "OE")
        s = unicodedata.normalize("NFKD", s)
        return "".join(ch for ch in s if not unicodedata.combining(ch))

    def _normalize_str(x) -> str:
        return ("" if x is None else str(x)).replace("\u00A0", " ").replace("\u202F", " ").strip()

    def _to_num_fr(x):
        if pd.isna(x):
            return np.nan
        s = str(x).replace("\u00A0", " ").replace("\u202F", " ").strip()
        if not s:
            return np.nan
        s = s.replace(" ", "").replace(",", ".")
        return pd.to_numeric(s, errors="coerce")

    def _as_bytes(x) -> bytes | None:
        if x is None:
            return None
        if isinstance(x, (bytes, bytearray)):
            return bytes(x)
        if isinstance(x, BytesIO):
            return x.getvalue()
        if isinstance(x, (str, os.PathLike, Path)):
            p = Path(x)
            if p.exists() and p.is_file():
                try:
                    return p.read_bytes()
                except Exception:
                    return None
        return None

    def _pick_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
        if df is None or df.empty:
            return None
        cols = {str(c).strip().lower(): c for c in df.columns}
        for cand in candidates:
            k = str(cand).strip().lower()
            if k in cols:
                return cols[k]
        for c in df.columns:
            lc = str(c).strip().lower()
            if any(str(cand).strip().lower() in lc for cand in candidates):
                return c
        return None

    def _normalize_vehicle_id_17(x: object) -> str:
        if x is None:
            return ""
        try:
            if isinstance(x, float) and pd.isna(x):
                return ""
        except Exception:
            pass

        s = str(x).strip()
        if not s or s.lower() in {"nan", "none"}:
            return ""

        if re.fullmatch(r"\d+\.0", s):
            s = s[:-2]

        s = s.replace("\u00A0", "").replace("\u202F", "").replace(" ", "")
        m = re.search(r"17-?(\d+)", s)
        if m:
            digits = m.group(1)
            return f"17-{digits.zfill(6)[-6:]}" if digits.isdigit() else ""
        if s.isdigit():
            return f"17-{s.zfill(6)[-6:]}"
        return ""

    # -----------------------------
    # Helpers used by Tab F
    # -----------------------------
    PLOTLY_CFG = {"displayModeBar": False, "responsive": True}

    def _layout_fig(fig, height: int = 380, showlegend: bool = False):
        return _layout_plotly_card(fig, height=height, title_pad=10, showlegend=showlegend)

    def card_open(title: str, subtitle: str = ""):
        _wrap_card_start(title, subtitle)

    def card_close():
        _wrap_card_end()

    def kpi_row(items: list[tuple[str, str]]):
        cols = st.columns(len(items), gap="large")
        for i, (label, val) in enumerate(items):
            with cols[i]:
                st.markdown(
                    f"""
                    <div style="padding:8px 10px;border:1px solid #eef0fb;border-radius:14px;">
                      <div class="card-sub" style="margin-bottom:6px;">{label}</div>
                      <div class="kpi" style="font-size:28px;">{val}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

    # -----------------------------
    # Load Global Vehicle List from brand workbooks (results)
    # -----------------------------
    def _find_vehicle_list_sheet(xls: pd.ExcelFile, brand: str):
        preferred = f"{brand} Vehicle List"
        if preferred in xls.sheet_names:
            return preferred
        for s in xls.sheet_names:
            if "vehicle list" in _strip_accents(s).lower():
                return s
        return None

    def load_global_vehicle_list_from_session(results_dict: dict) -> pd.DataFrame:
        rows = []
        for brand, res in results_dict.items():
            b = res.get("final_xlsx") or res.get("cleaned_xlsx") or res.get("final_bytes")
            b = _as_bytes(b)
            if not b:
                continue

            try:
                xls = pd.ExcelFile(BytesIO(b))
            except Exception:
                continue

            sheet = _find_vehicle_list_sheet(xls, brand)
            if not sheet:
                continue

            df = pd.read_excel(xls, sheet_name=sheet).rename(
                columns={
                    "Véhicule ID": "VehicleID",
                    "Vehicule ID": "VehicleID",
                    "Vehicle ID": "VehicleID",
                    "VehicleID": "VehicleID",
                    "Total HTVA": "TotalHTVA",
                    "Total": "TotalHTVA",
                    "TotalHTVA": "TotalHTVA",
                    "Brand(s)": "Brands",
                    "Brands": "Brands",
                }
            )

            if "VehicleID" not in df.columns or "TotalHTVA" not in df.columns:
                continue

            df["VehicleID"] = df["VehicleID"].apply(_normalize_str)
            df["TotalHTVA"] = df["TotalHTVA"].apply(_to_num_fr)

            df = df.dropna(subset=["TotalHTVA"]).copy()
            df = df[df["VehicleID"].ne("")].copy()
            df = df[~df["VehicleID"].str.contains("TOTAL", case=False, na=False)].copy()

            if "Brands" not in df.columns:
                df["Brands"] = str(brand)

            rows.append(df[["VehicleID", "TotalHTVA", "Brands"]].copy())

        if not rows:
            return pd.DataFrame(columns=["VehicleID", "TotalHTVA", "Brands"])

        all_df = pd.concat(rows, ignore_index=True)
        agg = (
            all_df.groupby("VehicleID", as_index=False)
            .agg(
                TotalHTVA=("TotalHTVA", "sum"),
                Brands=("Brands", lambda s: ", ".join(sorted(set(", ".join(map(str, s)).split(", "))))),
            )
            .copy()
        )
        agg["TotalHTVA"] = pd.to_numeric(agg["TotalHTVA"], errors="coerce").fillna(0.0).round(3)
        agg["Brands"] = agg["Brands"].astype(str).replace({"": "Unknown"}).fillna("Unknown")
        return agg

    # -----------------------------
    # Load Decompte from Dataset_Complet (Global Decompte Summary)
    # -----------------------------
    def _guess_global_dataset_bytes() -> bytes | None:
        b = _as_bytes(st.session_state.get("global_merge_path"))
        if b:
            return b
        return _as_bytes(st.session_state.get("global_merge_bytes"))

    def _pick_decompte_summary_sheet(sheet_names: list[str]) -> str | None:
        target = "global decompte summary"

        for s in sheet_names:
            if _strip_accents(s).strip().lower() == target:
                return s

        candidates = []
        for s in sheet_names:
            sl = _strip_accents(s).strip().lower()
            if sl.startswith(target):
                candidates.append(s)

        if not candidates:
            for s in sheet_names:
                sl = _strip_accents(s).strip().lower()
                if "decompte" in sl and "summary" in sl:
                    candidates.append(s)

        if not candidates:
            return None

        def suffix_num(name: str) -> int:
            m = re.search(r"_(\d+)$", str(name).strip())
            return int(m.group(1)) if m else 0

        return sorted(candidates, key=suffix_num, reverse=True)[0]

    def load_decompte_global_summary_from_dataset() -> pd.DataFrame:
        b = _guess_global_dataset_bytes()
        empty = pd.DataFrame(columns=["Brand", "TotalMain", "TotalPieces", "TotalFGB", "TotalDecompte"])
        if not b:
            return empty

        try:
            xls = pd.ExcelFile(BytesIO(b))
        except Exception:
            return empty

        sheet = _pick_decompte_summary_sheet(xls.sheet_names)
        if not sheet:
            return empty

        try:
            df = pd.read_excel(xls, sheet_name=sheet, header=2)
        except Exception:
            return empty

        if df.empty:
            return empty

        cols_norm = {_strip_accents(str(c)).strip().lower(): c for c in df.columns}

        def pick(*cands):
            for cand in cands:
                cn = _strip_accents(str(cand)).strip().lower()
                for k, orig in cols_norm.items():
                    if cn == k or cn in k:
                        return orig
            return None

        c_brand = pick("brand", "marque")
        c_main = pick("total main d'oeuvre", "total main d’œuvre", "total main", "main d'oeuvre", "main d oeuvre")
        c_piece = pick("total pieces", "total pièces", "total piece", "total pie", "pieces", "pièces")
        c_fgb = pick("fgb")
        c_tot = pick("total htva", "total htv/htva", "total htv", "total ht")

        if c_brand is None or c_tot is None:
            return empty

        out = pd.DataFrame(
            {
                "Brand": df[c_brand].astype(str).str.strip(),
                "TotalMain": df[c_main].apply(_to_num_fr) if (c_main and c_main in df.columns) else 0.0,
                "TotalPieces": df[c_piece].apply(_to_num_fr) if (c_piece and c_piece in df.columns) else 0.0,
                "TotalFGB": df[c_fgb].apply(_to_num_fr) if (c_fgb and c_fgb in df.columns) else 0.0,
                "TotalDecompte": df[c_tot].apply(_to_num_fr),
            }
        )

        out = out[~out["Brand"].str.contains("total", case=False, na=False)].copy()
        for c in ["TotalMain", "TotalPieces", "TotalFGB", "TotalDecompte"]:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)
        return out

    # -----------------------------
    # Tab 3: Vehicle ages
    # -----------------------------
    def load_vehicle_ages_from_session() -> pd.DataFrame:
        df_old = st.session_state.get("df_vehicle_dates_oldest")
        df_ear = st.session_state.get("df_vehicle_dates_earliest")

        df_age = None
        if isinstance(df_old, pd.DataFrame) and not df_old.empty:
            df_age = df_old
        elif isinstance(df_ear, pd.DataFrame) and not df_ear.empty:
            df_age = df_ear

        if isinstance(df_age, pd.DataFrame) and not df_age.empty:
            vcol = _pick_col(df_age, ["vehicle_raw", "vehicule", "véhicule", "vehicle"])
            acol = _pick_col(df_age, ["vehicle_age_years", "vehicle_age", "age", "years"])
            if vcol:
                out = pd.DataFrame()
                out["VehicleID"] = df_age[vcol].apply(_normalize_vehicle_id_17)
                out = out[out["VehicleID"] != ""].copy()
                out["vehicleAgeYears"] = pd.to_numeric(df_age[acol], errors="coerce") if acol else np.nan
                return out.drop_duplicates("VehicleID").reset_index(drop=True)

        b = _as_bytes(st.session_state.get("vehicle_dates_ds2"))
        if not b:
            return pd.DataFrame(columns=["VehicleID", "vehicleAgeYears"])

        try:
            df = pd.read_excel(BytesIO(b), sheet_name="earliest_per_vehicle")
        except Exception:
            return pd.DataFrame(columns=["VehicleID", "vehicleAgeYears"])

        vcol = _pick_col(df, ["vehicle_raw", "vehicule", "véhicule", "vehicle"])
        acol = _pick_col(df, ["vehicle_age_years", "vehicle_age", "age", "years"])
        if not vcol:
            return pd.DataFrame(columns=["VehicleID", "vehicleAgeYears"])

        out = pd.DataFrame()
        out["VehicleID"] = df[vcol].apply(_normalize_vehicle_id_17)
        out = out[out["VehicleID"] != ""].copy()
        out["vehicleAgeYears"] = pd.to_numeric(df[acol], errors="coerce") if acol else np.nan
        return out.drop_duplicates("VehicleID").reset_index(drop=True)

    # -----------------------------
    # Tab 3: Read counts from "Designation Stats"
    # -----------------------------
    def _norm_label(x) -> str:
        s = "" if x is None else str(x)
        s = s.strip().lower().replace("\u00a0", " ").replace("\u202f", " ")
        s = s.replace("œ", "oe").replace("’", "'")
        s = re.sub(r"\s+", " ", s)
        return s

    def _find_cell_contains(ws, text: str, max_rows: int = 250, max_cols: int = 80) -> tuple[int, int] | None:
        needle = _norm_label(text)
        for r in range(1, min(ws.max_row, max_rows) + 1):
            for c in range(1, min(ws.max_column, max_cols) + 1):
                if needle in _norm_label(ws.cell(r, c).value):
                    return (r, c)
        return None

    def _to_int_safe(x) -> int | None:
        if x is None:
            return None
        s = str(x).strip()
        if not s:
            return None
        if re.fullmatch(r"\d+\.0", s):
            s = s[:-2]
        try:
            return int(float(s))
        except Exception:
            return None

    def _read_designation_stats(ws) -> tuple[int | None, int | None]:
        anchor = _find_cell_contains(ws, "designation stats", max_rows=250, max_cols=80)
        if not anchor:
            return (None, None)

        ar, ac = anchor
        main_val = None
        piece_val = None

        for r in range(ar + 1, min(ar + 12, ws.max_row) + 1):
            label = _norm_label(ws.cell(r, ac).value)
            if not label:
                continue
            candidate = _to_int_safe(ws.cell(r, ac + 1).value)

            if ("main" in label) and ("oeuvre" in label or "d'oeuvre" in label or "d oeuvre" in label):
                if candidate is not None:
                    main_val = candidate

            if ("piece" in label) or ("pièce" in label) or ("pieces" in label) or ("pièces" in label):
                if candidate is not None:
                    piece_val = candidate

        if main_val is None or piece_val is None:
            for r in range(ar + 1, min(ar + 12, ws.max_row) + 1):
                for c in range(ac, min(ac + 6, ws.max_column) + 1):
                    t = _norm_label(ws.cell(r, c).value)
                    if not t:
                        continue
                    if main_val is None and ("main" in t) and ("oeuvre" in t or "d'oeuvre" in t or "d oeuvre" in t):
                        main_val = _to_int_safe(ws.cell(r, c + 1).value)
                    if piece_val is None and (("piece" in t) or ("pièce" in t) or ("pieces" in t) or ("pièces" in t)):
                        piece_val = _to_int_safe(ws.cell(r, c + 1).value)

        return (main_val, piece_val)

    def load_main_piece_visits_for_vehicles(ds_bytes: bytes, vehicle_ids: list[str]) -> pd.DataFrame:
        wb = load_workbook(BytesIO(ds_bytes), data_only=True)

        rows = []
        for vid in vehicle_ids:
            sheet_candidates = [f"V_{vid}", vid]
            sheet_name = next((s for s in sheet_candidates if s in wb.sheetnames), None)

            if not sheet_name:
                rows.append({"VehicleID": vid, "mainVisits": 0, "pieceVisits": 0, "totalVisits": 0, "sheetFound": False})
                continue

            ws = wb[sheet_name]
            main_stat, piece_stat = _read_designation_stats(ws)
            main_vis = int(main_stat or 0)
            piece_vis = int(piece_stat or 0)

            rows.append(
                {
                    "VehicleID": vid,
                    "mainVisits": main_vis,
                    "pieceVisits": piece_vis,
                    "totalVisits": main_vis + piece_vis,
                    "sheetFound": True,
                }
            )

        wb.close()
        out = pd.DataFrame(rows)
        for c in ["mainVisits", "pieceVisits", "totalVisits"]:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).astype(int)
        return out

    # -----------------------------
    # Designation df: from session OR dataset OR results (fallback)
    # -----------------------------
    def _get_designation_df_from_session() -> pd.DataFrame:
        for key in ["df_designation", "df_design", "designation_df", "df_designations", "df_designation_ops"]:
            v = st.session_state.get(key)
            if isinstance(v, pd.DataFrame) and not v.empty:
                return v.copy()
        return pd.DataFrame(columns=["Brand", "Type", "Designation", "Count"])

    def _build_designation_from_results(results_dict: dict) -> pd.DataFrame:
        # Build designation counts directly from cleaned dfs in results: df_main_clean / df_piece_clean
        def _pick_designation_col(df: pd.DataFrame):
            if df is None or df.empty:
                return None
            for c in df.columns:
                if _strip_accents(str(c)).strip().lower() in ("designation", "désignation"):
                    return c
            for c in df.columns:
                if "design" in _strip_accents(str(c)).lower():
                    return c
            return None

        def _clean_designation_raw(x) -> str:
            s = "" if x is None else str(x)
            s = s.replace("\u00A0", " ").replace("\u202F", " ").strip()
            s = s.replace("’", "'")
            s = re.sub(r"\s+", " ", s)
            return s

        def _designation_key(x) -> str:
            s = _clean_designation_raw(x).lower()
            s = _strip_accents(s)
            s = re.sub(r"[^\w\s'-]", " ", s)
            s = re.sub(r"\s+", " ", s).strip()
            return s

        def _compute(df_clean: pd.DataFrame, brand: str, kind: str) -> pd.DataFrame:
            if not isinstance(df_clean, pd.DataFrame) or df_clean.empty:
                return pd.DataFrame(columns=["Brand", "Type", "Designation", "Count"])
            des_col = _pick_designation_col(df_clean)
            if not des_col:
                return pd.DataFrame(columns=["Brand", "Type", "Designation", "Count"])

            tmp = pd.DataFrame({"DesignationRaw": df_clean[des_col]})
            tmp["Key"] = tmp["DesignationRaw"].apply(_designation_key)
            tmp["Designation"] = tmp["DesignationRaw"].apply(_clean_designation_raw)
            tmp = tmp[tmp["Key"].ne("")].copy()

            if tmp.empty:
                return pd.DataFrame(columns=["Brand", "Type", "Designation", "Count"])

            out = (
                tmp.groupby("Key", as_index=False)
                .agg(
                    Count=("Key", "size"),
                    Designation=("Designation", lambda s: s.mode().iloc[0] if not s.mode().empty else s.iloc[0]),
                )
                .sort_values("Count", ascending=False)
                .reset_index(drop=True)
            )
            out.insert(0, "Type", kind)
            out.insert(0, "Brand", brand)
            out["Count"] = out["Count"].astype(int)
            return out[["Brand", "Type", "Designation", "Count"]]

        all_rows = []
        for brand, res in (results_dict or {}).items():
            all_rows.append(_compute(res.get("df_main_clean"), str(brand), "Main d'œuvre"))
            all_rows.append(_compute(res.get("df_piece_clean"), str(brand), "Pièce"))

        if not all_rows:
            return pd.DataFrame(columns=["Brand", "Type", "Designation", "Count"])

        df = pd.concat(all_rows, ignore_index=True)
        df["Designation"] = df["Designation"].astype(str).str.strip()
        df["Count"] = pd.to_numeric(df["Count"], errors="coerce").fillna(0).astype(int)
        df = df[df["Designation"].ne("")].copy()
        return df

    # -----------------------------
    # Build datasets
    # -----------------------------
    dfv = load_global_vehicle_list_from_session(results)
    if dfv.empty:
        st.error("Could not build Global Vehicle List from brand workbooks.")
        return

    df_decompte = load_decompte_global_summary_from_dataset()
    dataset_bytes = _guess_global_dataset_bytes()
    df_ages = load_vehicle_ages_from_session()

    # df_design: session -> dataset hyperlinks -> results fallback
    df_design = _get_designation_df_from_session()
    if (df_design is None or df_design.empty) and dataset_bytes:
        df_design = build_df_design_from_dataset(dataset_bytes, _pick_col)
    if df_design is None or df_design.empty:
        df_design = _build_designation_from_results(results)
    st.session_state["df_designation"] = df_design  # cache

    brand_global = st.session_state.get("brand_global", "All")

    # =========================================================
    # Tabs
    # =========================================================
    tab_overview, tab_brand, tab_corr, tabF = st.tabs(
        [
            "Overview",
            "Brand Summary (Vehicles / FGB / Décompte)",
            "Age vs Main/Pieces Times Done vs Cost",
            "Brand Operations (Pièces & Main d'œuvre)",
        ]
    )

    # =========================================================
    # OVERVIEW TAB
    # =========================================================
    with tab_overview:
        total_cost = float(dfv["TotalHTVA"].sum())
        n_vehicles = int(dfv["VehicleID"].nunique())
        avg_cost = float(dfv["TotalHTVA"].mean()) if len(dfv) else 0.0
        median_cost = float(dfv["TotalHTVA"].median()) if len(dfv) else 0.0
        min_cost = float(dfv["TotalHTVA"].min()) if len(dfv) else 0.0
        max_cost = float(dfv["TotalHTVA"].max()) if len(dfv) else 0.0
        cost_range = max_cost - min_cost

        dfv_sorted = dfv.sort_values("TotalHTVA", ascending=False).reset_index(drop=True)
        top10 = dfv_sorted.head(10)
        top10_pct = (float(top10["TotalHTVA"].sum()) / total_cost * 100) if total_cost else 0.0

        decompte_total_htva = float(df_decompte["TotalDecompte"].sum()) if not df_decompte.empty else 0.0
        decompte_total_fgb = float(df_decompte["TotalFGB"].sum()) if not df_decompte.empty else 0.0
        recon_diff = total_cost - decompte_total_htva

        st.subheader("Overview (Global)")

        r1 = st.columns(4, gap="large")
        with r1[0]:
            card_kpi("Total Vehicle HTVA", f"{total_cost:,.3f} {CURRENCY}", "Global", "blue")
        with r1[1]:
            card_kpi("Total Vehicles", f"{n_vehicles:,}", "Global", "muted")
        with r1[2]:
            card_kpi("Avg Cost / Vehicle", f"{avg_cost:,.3f} {CURRENCY}", "Global", "pink")
        with r1[3]:
            card_kpi("Median Cost / Vehicle", f"{median_cost:,.3f} {CURRENCY}", "Global", "muted")

        r2 = st.columns(4, gap="large")
        with r2[0]:
            card_kpi("% Cost from Top 10", f"{top10_pct:.1f}%", "Global", "muted")
        with r2[1]:
            card_kpi("Max Vehicle HTVA", f"{max_cost:,.3f} {CURRENCY}", "Global", "pink")
        with r2[2]:
            card_kpi("Range (Max–Min)", f"{cost_range:,.3f} {CURRENCY}", "Global", "muted")
        with r2[3]:
            card_kpi("Min Vehicle HTVA", f"{min_cost:,.3f} {CURRENCY}", "Global", "muted")

        st.markdown("##")
        st.subheader("Scope & Reconciliation")

        r3 = st.columns(4, gap="large")
        with r3[0]:
            card_kpi("Vehicle HTVA Total", f"{total_cost:,.3f} {CURRENCY}", "Global", "blue")
        with r3[1]:
            card_kpi(
                "Decompte Total (HTVA/HTV)",
                f"{decompte_total_htva:,.3f} {CURRENCY}" if decompte_total_htva > 0 else "n/a",
                "Global merge",
                "muted",
            )
        with r3[2]:
            card_kpi(
                "Difference (Vehicle − Decompte)",
                f"{recon_diff:,.3f} {CURRENCY}" if decompte_total_htva > 0 else "n/a",
                "Global merge",
                "pink" if (decompte_total_htva > 0 and abs(recon_diff) > 0) else "muted",
            )
        with r3[3]:
            card_kpi(
                "FGB (sum)",
                f"{decompte_total_fgb:,.3f} {CURRENCY}" if decompte_total_htva > 0 else "n/a",
                "Global merge",
                "green",
            )

        st.markdown("##")

        def split_brands(s: str) -> list[str]:
            parts = [p.strip() for p in str(s).split(",")]
            parts = [p for p in parts if p]
            return parts if parts else ["Unknown"]

        brand_rows = dfv.copy()
        brand_rows["BrandList"] = brand_rows["Brands"].apply(split_brands)
        brand_rows["nBrands"] = brand_rows["BrandList"].apply(len).replace(0, 1)
        brand_rows = brand_rows.explode("BrandList")
        brand_rows["Brand"] = brand_rows["BrandList"].astype(str).str.strip()
        brand_rows["AllocatedCost"] = brand_rows["TotalHTVA"] / brand_rows["nBrands"]

        brand_agg = (
            brand_rows.groupby("Brand", as_index=False)
            .agg(TotalHTVA=("AllocatedCost", "sum"))
            .sort_values("TotalHTVA", ascending=False)
            .reset_index(drop=True)
        )
        brand_agg["TotalHTVA"] = pd.to_numeric(brand_agg["TotalHTVA"], errors="coerce").fillna(0.0).round(3)

        left, right = st.columns([1.6, 1.0], gap="large")
        with left:
            _wrap_card_start("Total HTVA by Brand", CURRENCY)
            fig_brand = px.bar(brand_agg, x="Brand", y="TotalHTVA", color_discrete_sequence=[BLUE])
            _layout_plotly_card(fig_brand, height=420, title_pad=10, showlegend=False)
            st.plotly_chart(fig_brand, use_container_width=True, config=PLOTLY_CFG)
            _wrap_card_end()

        with right:
            _wrap_card_start("Brand Share of Total HTVA (%)", "")
            share_df = brand_agg.copy()
            denom = float(share_df["TotalHTVA"].sum()) if not share_df.empty else 0.0
            share_df["SharePct"] = (share_df["TotalHTVA"] / denom * 100) if denom else 0.0
            share_df["SharePct"] = pd.to_numeric(share_df["SharePct"], errors="coerce").fillna(0.0)

            fig_share = px.bar(
                share_df.sort_values("SharePct", ascending=True),
                x="SharePct",
                y="Brand",
                orientation="h",
                text="SharePct",
                color_discrete_sequence=[PINK],
            )
            fig_share.update_traces(texttemplate="%{text:.1f}%", textposition="outside", cliponaxis=False)
            fig_share.update_xaxes(title="Share (%)")
            fig_share.update_yaxes(title="")
            _layout_plotly_card(fig_share, height=420, title_pad=10, showlegend=False)
            st.plotly_chart(fig_share, use_container_width=True, config=PLOTLY_CFG)
            _wrap_card_end()

        st.markdown("##")
        b1, b2 = st.columns([1.1, 1.0], gap="large")

        BINS_FIXED = [0, 500, 1000, 2500, 5000, np.inf]
        BIN_LABELS_FIXED = ["0–500", "500–1000", "1000–2500", "2500–5000", "5000+"]

        dfv_plot = dfv.copy()
        dfv_plot["CostBucket"] = pd.cut(
            dfv_plot["TotalHTVA"],
            bins=BINS_FIXED,
            labels=BIN_LABELS_FIXED,
            include_lowest=True,
            right=False,
        )

        dist_agg = dfv_plot.groupby("CostBucket", as_index=False, observed=False).agg(
            Vehicles=("VehicleID", "nunique"),
            TotalCost=("TotalHTVA", "sum"),
        )

        with b1:
            _wrap_card_start("Cost Distribution (Vehicle count)", "")
            fig_dist = px.bar(dist_agg, x="CostBucket", y="Vehicles", color_discrete_sequence=[GREEN])
            fig_dist.update_xaxes(title="")
            fig_dist.update_yaxes(title="Vehicles")
            _layout_plotly_card(fig_dist, height=340, title_pad=10, showlegend=False)
            st.plotly_chart(fig_dist, use_container_width=True, config=PLOTLY_CFG)
            _wrap_card_end()

        with b2:
            _wrap_card_start("Top 10 Most Expensive Vehicles", "")
            st.dataframe(top10[["VehicleID", "TotalHTVA", "Brands"]], use_container_width=True, height=340)
            _wrap_card_end()

    # =========================================================
    # BRAND SUMMARY TAB
    # =========================================================
    with tab_brand:
        st.subheader("Brands Summary")

        def split_brands(s: str) -> list[str]:
            parts = [p.strip() for p in str(s).split(",")]
            parts = [p for p in parts if p]
            return parts if parts else ["Unknown"]

        brand_rows = dfv.copy()
        brand_rows["BrandList"] = brand_rows["Brands"].apply(split_brands)
        brand_rows["nBrands"] = brand_rows["BrandList"].apply(len).replace(0, 1)
        brand_rows = brand_rows.explode("BrandList")
        brand_rows["Brand"] = brand_rows["BrandList"].astype(str).str.strip()
        brand_rows["VehicleHTVAAllocated"] = brand_rows["TotalHTVA"] / brand_rows["nBrands"]

        veh_alloc = (
            brand_rows.groupby("Brand", as_index=False)
            .agg(
                Vehicles=("VehicleID", "nunique"),
                VehicleHTVAAllocated=("VehicleHTVAAllocated", "sum"),
            )
            .copy()
        )

        if isinstance(df_decompte, pd.DataFrame) and not df_decompte.empty and "Brand" in df_decompte.columns:
            dec = df_decompte.copy()
            dec["Brand"] = dec["Brand"].astype(str).str.strip()
            for c in ["TotalDecompte", "TotalFGB", "TotalMain", "TotalPieces"]:
                if c in dec.columns:
                    dec[c] = pd.to_numeric(dec[c], errors="coerce").fillna(0.0)
                else:
                    dec[c] = 0.0
            dec = dec.groupby("Brand", as_index=False).agg(
                TotalDecompte=("TotalDecompte", "sum"),
                TotalFGB=("TotalFGB", "sum"),
                TotalMain=("TotalMain", "sum"),
                TotalPieces=("TotalPieces", "sum"),
            )
        else:
            dec = pd.DataFrame(columns=["Brand", "TotalDecompte", "TotalFGB", "TotalMain", "TotalPieces"])

        summary = veh_alloc.merge(dec, on="Brand", how="left")

        for c in ["TotalDecompte", "TotalFGB", "TotalMain", "TotalPieces"]:
            if c not in summary.columns:
                summary[c] = 0.0
            summary[c] = pd.to_numeric(summary[c], errors="coerce").fillna(0.0)

        summary["VehicleHTVAAllocated"] = pd.to_numeric(summary["VehicleHTVAAllocated"], errors="coerce").fillna(0.0)
        summary["Vehicles"] = pd.to_numeric(summary["Vehicles"], errors="coerce").fillna(0).astype(int)
        summary["Difference (Vehicle − Decompte)"] = summary["VehicleHTVAAllocated"] - summary["TotalDecompte"]
        summary = summary.sort_values("VehicleHTVAAllocated", ascending=False).reset_index(drop=True)

        k = st.columns(4, gap="large")
        with k[0]:
            card_kpi("Brands", f"{summary['Brand'].nunique():,}", "muted")
        with k[1]:
            card_kpi("Vehicle HTVA", f"{summary['VehicleHTVAAllocated'].sum():,.3f} {CURRENCY}", "blue")

        total_dec = float(summary["TotalDecompte"].sum())
        total_fgb = float(summary["TotalFGB"].sum())
        with k[2]:
            card_kpi("Décompte Total", f"{total_dec:,.3f} {CURRENCY}" if total_dec > 0 else "n/a", "muted")
        with k[3]:
            card_kpi("FGB Total", f"{total_fgb:,.3f} {CURRENCY}" if total_dec > 0 else "n/a", "green")

        st.markdown("##")
        _wrap_card_start("Brands table")

        compact_cols = [
            "Brand",
            "Vehicles",
            "VehicleHTVAAllocated",
            "TotalDecompte",
            "Difference (Vehicle − Decompte)",
            "TotalFGB",
        ]
        compact_cols = [c for c in compact_cols if c in summary.columns]

        tbl = summary[compact_cols].copy()
        for c in tbl.columns:
            if c != "Brand":
                tbl[c] = pd.to_numeric(tbl[c], errors="coerce")

        st.dataframe(tbl, use_container_width=True, hide_index=True, height=280)
        _wrap_card_end()

        if total_dec > 0:
            st.markdown("##")
            _wrap_card_start("Vehicle vs Décompte by Brand", "")
            chart_df = summary[["Brand", "VehicleHTVAAllocated", "TotalDecompte"]].copy()
            chart_df = chart_df.melt(id_vars="Brand", var_name="Metric", value_name="Amount")
            # enforce blue vs pink for the two metrics
            fig = px.bar(
                chart_df,
                x="Brand",
                y="Amount",
                color="Metric",
                barmode="group",
                color_discrete_map={
                    "VehicleHTVAAllocated": BLUE,
                    "TotalDecompte": PINK,
                },
            )
            _layout_plotly_card(fig, height=380, title_pad=10, showlegend=True)
            st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CFG)
            _wrap_card_end()

    # =========================================================
    # TAB 3: Age vs Main/Pieces Times Done vs Cost
    # =========================================================
    with tab_corr:
        st.subheader("(Age Vehicle) vs (Designations 'Main d'oeuvre/Pièces') vs Cost")

        if not dataset_bytes:
            st.info("Dataset_Complet is not loaded in this session. Build Global Merge first.")
            st.stop()

        base = dfv.sort_values("TotalHTVA", ascending=False).copy()
        base["VehicleID"] = base["VehicleID"].apply(_normalize_vehicle_id_17)

        corr = base.copy()
        if isinstance(df_ages, pd.DataFrame) and not df_ages.empty:
            df_ages2 = df_ages.copy()
            df_ages2["VehicleID"] = df_ages2["VehicleID"].apply(_normalize_vehicle_id_17)
            corr = corr.merge(df_ages2[["VehicleID", "vehicleAgeYears"]], on="VehicleID", how="left")
        else:
            corr["vehicleAgeYears"] = np.nan

        corr["vehicleAgeYears"] = pd.to_numeric(corr["vehicleAgeYears"], errors="coerce")

        vids = corr["VehicleID"].dropna().astype(str).tolist()
        with st.spinner("Reading Designation Stats (Main/Pièces) per vehicle..."):
            df_vis = load_main_piece_visits_for_vehicles(dataset_bytes, vids)

        corr = corr.merge(df_vis, on="VehicleID", how="left")

        k = st.columns(4, gap="large")
        with k[0]:
            card_kpi("Vehicles analyzed", f"{len(corr):,}",  "muted")
        with k[1]:
            avg_age = corr["vehicleAgeYears"].dropna().mean() if corr["vehicleAgeYears"].notna().any() else np.nan
            card_kpi("Avg age (years)", f"{avg_age:.1f}" if pd.notna(avg_age) else "n/a",  "muted")
        with k[2]:
            card_kpi("Avg Main times", f"{pd.to_numeric(corr['mainVisits'], errors='coerce').fillna(0).mean():.1f}", "pink")
        with k[3]:
            card_kpi("Avg Pièces times", f"{pd.to_numeric(corr['pieceVisits'], errors='coerce').fillna(0).mean():.1f}",  "green")

        st.markdown("##")

        corr2 = corr.dropna(subset=["vehicleAgeYears"]).copy()
        if corr2.empty:
            st.info("No ages found. Run Vehicle Dates Extraction first (earliest_per_vehicle).")
            st.stop()

        _wrap_card_start("Correlation Matrix (Age vs Fixing vs Spending)")

        tmp = corr2.copy()
        tmp["AgeYears"] = pd.to_numeric(tmp["vehicleAgeYears"], errors="coerce")
        tmp["TotalHTVA"] = pd.to_numeric(tmp["TotalHTVA"], errors="coerce")
        tmp["mainVisits"] = pd.to_numeric(tmp.get("mainVisits", np.nan), errors="coerce")
        tmp["pieceVisits"] = pd.to_numeric(tmp.get("pieceVisits", np.nan), errors="coerce")
        tmp["totalVisits"] = pd.to_numeric(tmp.get("totalVisits", np.nan), errors="coerce")

        cols = ["AgeYears", "mainVisits", "pieceVisits", "totalVisits", "TotalHTVA"]
        tmp = tmp[cols].dropna(how="all")
        corr_mat = tmp.corr(numeric_only=True).round(3)

        pink_blue_green = [
            (0.00, "#FBCFE8"),
            (0.50, "#BFDBFE"),
            (1.00, "#BBF7D0"),
        ]

        fig = px.imshow(
            corr_mat,
            text_auto=True,
            zmin=-1,
            zmax=1,
            color_continuous_scale=pink_blue_green,
            aspect="auto",
        )
        fig.update_layout(
            height=520,
            margin=dict(l=10, r=10, t=10, b=10),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            coloraxis_colorbar=dict(title="Corr", tickvals=[-1, -0.5, 0, 0.5, 1]),
        )
        fig.update_traces(textfont_size=12)
        _fig_black_text(fig)

        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CFG)
        _wrap_card_end()

        st.markdown("##")

        bins = [-1, 5, 10, 15, 20, 25, 1000]
        labels = ["0–5", "6–10", "11–15", "16–20", "21–25", "26+"]

        corr2["AgeGroup"] = pd.cut(corr2["vehicleAgeYears"], bins=bins, labels=labels)

        grp = (
            corr2.groupby("AgeGroup", as_index=False)
            .agg(
                Vehicles=("VehicleID", "nunique"),
                AvgHTVA=("TotalHTVA", "mean"),
                AvgMainTimes=("mainVisits", "mean"),
                AvgPieceTimes=("pieceVisits", "mean"),
            )
            .copy()
        )
        for c in ["AvgHTVA", "AvgMainTimes", "AvgPieceTimes"]:
            grp[c] = pd.to_numeric(grp[c], errors="coerce").fillna(0.0)

        grp["AvgHTVA"] = grp["AvgHTVA"].round(3)
        grp["AvgMainTimes"] = grp["AvgMainTimes"].round(2)
        grp["AvgPieceTimes"] = grp["AvgPieceTimes"].round(2)

        plot_grp = grp[grp["Vehicles"] > 0].copy()

        _wrap_card_start("Average Cost (HTVA) by Age Group")

        donut_seq = [BLUE, PINK, GREEN, BLUE_L, PINK_L, GREEN_L]
        fig_a = px.pie(
            plot_grp,
            names="AgeGroup",
            values="AvgHTVA",
            hole=0.45,
            color="AgeGroup",
            color_discrete_sequence=donut_seq,
        )
        fig_a.update_traces(textinfo="percent+label")
        _layout_plotly_card(fig_a, height=360, title_pad=10, showlegend=False)
        st.plotly_chart(fig_a, use_container_width=True, config=PLOTLY_CFG)
        _wrap_card_end()

        st.markdown("##")

        _wrap_card_start("Average Main d'œuvre vs Pièces Times Done by Age Group")
        combo = grp[["AgeGroup", "AvgMainTimes", "AvgPieceTimes"]].copy()
        combo = combo.melt(
            id_vars="AgeGroup",
            value_vars=["AvgMainTimes", "AvgPieceTimes"],
            var_name="Type",
            value_name="AvgTimes",
        )
        combo["Type"] = combo["Type"].replace({"AvgMainTimes": "Main d'œuvre", "AvgPieceTimes": "Pièces"})

        fig = px.bar(
            combo,
            x="AgeGroup",
            y="AvgTimes",
            color="Type",
            barmode="group",
            text="AvgTimes",
            color_discrete_map={"Main d'œuvre": PINK, "Pièces": GREEN},
        )
        fig.update_traces(texttemplate="%{text:.2f}", textposition="outside", cliponaxis=False)
        fig.update_yaxes(title="Avg times")
        fig.update_xaxes(title="Age group")
        _layout_plotly_card(fig, height=380, title_pad=10, showlegend=True)
        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CFG)
        _wrap_card_end()

        st.markdown("##")

        _wrap_card_start("Vehicle Maintenance & Cost Summary", "Age + Main d'oeuvre/Pièces + HTVA")
        table = corr.copy()
        table["vehicleAgeYears"] = pd.to_numeric(table["vehicleAgeYears"], errors="coerce")
        table = table.rename(columns={"vehicleAgeYears": "AgeYears"})

        cols = ["VehicleID", "Brands", "TotalHTVA", "AgeYears", "mainVisits", "pieceVisits", "totalVisits"]
        cols = [c for c in cols if c in table.columns]

        st.dataframe(table[cols].sort_values("TotalHTVA", ascending=False), use_container_width=True, height=520)
        _wrap_card_end()

    # ------------------------------------------------------------
    # Tab F — Brand Operations
    # ------------------------------------------------------------
    with tabF:
        st.subheader("Brand Operations Dashboard (Pièces & Main d'œuvre)")

        if df_design.empty:
            st.info("No designation data available.")
            return

        brands_ops = sorted(df_design["Brand"].unique().tolist())
        if brand_global != "All" and brand_global in brands_ops:
            brand_choice = brand_global
            st.caption(f"Brand locked by global filter: **{brand_global}**")
        else:
            brand_choice = st.selectbox("Brand selector", brands_ops, key="ops_brand_core")

        df_b = df_design[df_design["Brand"] == brand_choice].copy()

        left, right = st.columns(2, gap="large")

        with left:
            df_piece = df_b[df_b["Type"] == "Pièce"].copy()
            total_pieces = int(df_piece["Count"].sum()) if not df_piece.empty else 0
            distinct_parts = int(df_piece["Designation"].nunique()) if not df_piece.empty else 0

            top10_piece = (
                df_piece.groupby("Designation", as_index=False)
                .agg(Count=("Count", "sum"))
                .sort_values("Count", ascending=False)
                .head(10)
            )

            card_open("Pièces", f"Brand: {brand_choice}")
            kpi_row([("Total Pieces", f"{total_pieces:,}"), ("Distinct Parts", f"{distinct_parts:,}")])

            if not top10_piece.empty:
                fig = px.bar(
                    top10_piece.sort_values("Count", ascending=True),
                    x="Count",
                    y="Designation",
                    orientation="h",
                    color_discrete_sequence=[GREEN],
                )
                _layout_fig(fig, height=380, showlegend=False)
                st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CFG)
            else:
                st.info("No part designations found for this brand.")
            card_close()

        with right:
            df_main = df_b[df_b["Type"] == "Main d'œuvre"].copy()
            total_interventions = int(df_main["Count"].sum()) if not df_main.empty else 0
            distinct_ops = int(df_main["Designation"].nunique()) if not df_main.empty else 0

            top10_main = (
                df_main.groupby("Designation", as_index=False)
                .agg(Count=("Count", "sum"))
                .sort_values("Count", ascending=False)
                .head(10)
            )

            card_open("Main d'œuvre", f"Brand: {brand_choice}")
            kpi_row([("Total Interventions", f"{total_interventions:,}"), ("Distinct Operations", f"{distinct_ops:,}")])

            if not top10_main.empty:
                fig = px.bar(
                    top10_main.sort_values("Count", ascending=True),
                    x="Count",
                    y="Designation",
                    orientation="h",
                    color_discrete_sequence=[PINK],
                )
                _layout_fig(fig, height=380, showlegend=False)
                st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CFG)
            else:
                st.info("No labor designations found for this brand.")
            card_close()