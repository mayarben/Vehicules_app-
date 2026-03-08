# utils/vehicle_dates.py
from __future__ import annotations

import os
import re
import math
from io import BytesIO
from datetime import datetime, date, timedelta

from openpyxl import load_workbook, Workbook
from openpyxl.styles.numbers import is_date_format


# =======================
# RAW / AS-SEEN VALUE
# =======================
def cell_display_value(cell):
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    if isinstance(v, datetime):
        if (
            v.time().hour == 0
            and v.time().minute == 0
            and v.time().second == 0
            and v.time().microsecond == 0
        ):
            return v.strftime("%d/%m/%Y")
        return v.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if is_date_format(cell.number_format) and 20000 < v < 60000:
            base = datetime(1899, 12, 30)
            dt = base + timedelta(days=v)
            return dt.strftime("%d/%m/%Y")
        if math.isfinite(v) and abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return str(v)
    return str(v)


# =======================
# PARSE DATE + AGE
# =======================
def parse_date_raw(s: str):
    if s is None:
        return None
    s2 = str(s).strip()
    if not s2:
        return None

    m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", s2)
    if not m:
        return None

    d = int(m.group(1))
    mo = int(m.group(2))
    y = int(m.group(3))
    if y < 100:
        y += 2000  # adjust if needed

    try:
        return date(y, mo, d)
    except ValueError:
        return None


def age_years_from(dt: date):
    if dt is None:
        return ""
    today = date.today()
    years = today.year - dt.year
    if (today.month, today.day) < (dt.month, dt.day):
        years -= 1
    return str(years)


# =======================
# VEHICLE/DATE DETECTION
# =======================
def normalize(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


vehicle_header_keywords = [
    "vehicule",
    "véhicule",
    "vehicle",
    "immat",
    "immatriculation",
    "matricule",
    "veh",
    "vh",
]
date_header_keywords = [
    "date",
    "dt",
    "date entrée",
    "date entree",
    "date d'entrée",
    "date sortie",
    "entree",
    "sortie",
]


def find_columns_by_headers(ws, max_rows=150, max_cols=120):
    veh_hits, date_hits = [], []
    for r in range(1, min(max_rows, ws.max_row) + 1):
        for c in range(1, min(max_cols, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            txt = normalize(v)
            if not txt:
                continue
            if any(k in txt for k in vehicle_header_keywords):
                veh_hits.append((r, c))
            if any(k in txt for k in date_header_keywords):
                date_hits.append((r, c))

    if not veh_hits or not date_hits:
        return None, None, None

    best = None
    for (rv, cv) in veh_hits:
        for (rd, cd) in date_hits:
            row_bonus = 100 if rv == rd else -abs(rv - rd)
            dist_penalty = abs(cv - cd)
            score = row_bonus - dist_penalty
            header_row = rv if rv == rd else min(rv, rd)
            if best is None or score > best[0]:
                best = (score, header_row, cv, cd)

    _, header_row, veh_col, date_col = best
    return veh_col, date_col, header_row


def looks_like_date_value(v, cell):
    if v is None:
        return False
    if isinstance(v, (datetime, date)):
        return True
    if isinstance(v, str):
        return bool(re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", v.strip()))
    if isinstance(v, float) and is_date_format(cell.number_format):
        return True
    return False


def vehicle_score(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        if isinstance(v, float) and abs(v - round(v)) > 1e-9:
            return 0
        iv = int(round(v))
        return 10 if 1000 <= iv <= 99999999 else 0
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return 0
        if s.lower() in {"voiture", "véhicule", "vehicule"}:
            return 0
        return 12 if re.fullmatch(r"\d{4,10}", s) else 0
    return 0


def infer_columns_by_scoring(ws, scan_rows=250, scan_cols=60):
    max_r = min(ws.max_row, scan_rows)
    max_c = min(ws.max_column, scan_cols)

    veh_scores = [0] * (max_c + 1)
    date_scores = [0] * (max_c + 1)

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(r, c)
            veh_scores[c] += vehicle_score(cell.value)
            if looks_like_date_value(cell.value, cell):
                date_scores[c] += 4

    veh_col = max(range(1, max_c + 1), key=lambda c: veh_scores[c], default=None)
    date_col = max(range(1, max_c + 1), key=lambda c: date_scores[c], default=None)

    if (
        veh_col is None
        or date_col is None
        or veh_scores[veh_col] == 0
        or date_scores[date_col] == 0
    ):
        return None, None, None

    first_data_row = 1
    for r in range(1, max_r + 1):
        if vehicle_score(ws.cell(r, veh_col).value) >= 10:
            first_data_row = r
            break

    return veh_col, date_col, max(0, first_data_row - 1)


def extract_from_sheet(ws):
    veh_col, date_col, header_row = find_columns_by_headers(ws)
    if veh_col is None or date_col is None:
        veh_col, date_col, header_row = infer_columns_by_scoring(ws)

    if veh_col is None or date_col is None:
        return []

    start_row = (header_row or 0) + 1
    rows = []
    empty_run = 0

    for r in range(start_row, ws.max_row + 1):
        veh_raw = cell_display_value(ws.cell(r, veh_col))
        date_raw = cell_display_value(ws.cell(r, date_col))

        if veh_raw == "":
            empty_run += 1
            if empty_run >= 30:
                break
            continue
        empty_run = 0

        if not re.fullmatch(r"\d{4,10}", str(veh_raw).strip()):
            continue

        rows.append((veh_raw, date_raw, r))
    return rows


def run_vehicle_date_extraction(input_files: list[str]) -> tuple[bytes, bytes]:
    """
    Returns:
      dataset1_bytes -> vehicle_date_extraction.xlsx (all rows + earliest date per vehicle)
      dataset2_bytes -> vehicle_earliest_only.xlsx (one row per vehicle)
    """
    all_rows = []
    earliest_date_by_vehicle = {}

    first_row = {}
    earliest_row = {}

    for fp in input_files:
        if not os.path.exists(fp):
            continue

        wb = load_workbook(fp, data_only=False)
        for sh in wb.sheetnames:
            ws = wb[sh]
            extracted = extract_from_sheet(ws)

            for veh_raw, date_raw, excel_row in extracted:
                src = os.path.basename(fp)

                all_rows.append((src, sh, excel_row, veh_raw, date_raw))

                if veh_raw not in first_row:
                    first_row[veh_raw] = (src, sh, excel_row, date_raw)

                dt = parse_date_raw(date_raw)
                if dt is not None:
                    cur = earliest_date_by_vehicle.get(veh_raw)
                    if cur is None or dt < cur:
                        earliest_date_by_vehicle[veh_raw] = dt

                    cur2 = earliest_row.get(veh_raw)
                    if cur2 is None or dt < cur2[0]:
                        earliest_row[veh_raw] = (dt, src, sh, excel_row, date_raw)

        wb.close()

    # ---------- Dataset 1 ----------
    out_wb1 = Workbook()
    ws_out1 = out_wb1.active
    ws_out1.title = "extraction"
    ws_out1.append(
        [
            "source_file",
            "sheet",
            "excel_row",
            "vehicle_raw",  # NO 17- HERE
            "date_raw",
            "vehicle_earliest_date",
            "vehicle_age_years",
        ]
    )
    ws_out1.freeze_panes = "A2"

    ws_map = out_wb1.create_sheet("VEHICLE_MIN_DATE")
    ws_map.append(["vehicle_raw", "earliest_date", "age_years"])  # NO 17- HERE
    for veh_raw, dt in sorted(earliest_date_by_vehicle.items(), key=lambda x: x[0]):
        ws_map.append([veh_raw, dt.strftime("%d/%m/%Y"), age_years_from(dt)])

    for src, sh, excel_row, veh_raw, date_raw in all_rows:
        dt_earliest = earliest_date_by_vehicle.get(veh_raw)
        earliest_str = dt_earliest.strftime("%d/%m/%Y") if dt_earliest else ""
        age_str = age_years_from(dt_earliest) if dt_earliest else ""
        ws_out1.append([src, sh, excel_row, veh_raw, date_raw, earliest_str, age_str])

    for col, w in zip(["A", "B", "C", "D", "E", "F", "G"], [28, 26, 10, 22, 22, 20, 18]):
        ws_out1.column_dimensions[col].width = w

    # ---------- Dataset 2 ----------
    out_wb2 = Workbook()
    ws_out2 = out_wb2.active
    ws_out2.title = "earliest_per_vehicle"
    ws_out2.append(
        [
            "source_file",
            "sheet",
            "excel_row",
            "vehicle_raw",  # KEEP 17- HERE
            "date_raw",
            "vehicle_earliest_date",
            "vehicle_age_years",
        ]
    )
    ws_out2.freeze_panes = "A2"

    def sort_key(v):
        s = str(v).strip()
        if s.isdigit():
            return (0, s.zfill(12))
        return (1, s.lower())

    for veh_raw in sorted(first_row.keys(), key=sort_key):
        if veh_raw in earliest_row:
            dt, src, sh, excel_row, date_raw = earliest_row[veh_raw]
            earliest_str = dt.strftime("%d/%m/%Y")
            age_str = age_years_from(dt)
        else:
            src, sh, excel_row, date_raw = first_row[veh_raw]
            earliest_str = ""
            age_str = ""

        veh_raw_prefixed = f"17-{veh_raw}" if str(veh_raw).strip() else ""
        ws_out2.append([src, sh, excel_row, veh_raw_prefixed, date_raw, earliest_str, age_str])

    for col, w in zip(["A", "B", "C", "D", "E", "F", "G"], [28, 26, 10, 22, 22, 20, 18]):
        ws_out2.column_dimensions[col].width = w

    bio1 = BytesIO()
    out_wb1.save(bio1)

    bio2 = BytesIO()
    out_wb2.save(bio2)

    return bio1.getvalue(), bio2.getvalue()