import re
from io import BytesIO
from datetime import datetime

import streamlit as st
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.ui import inject_css, hero, card_open, card_close
from utils.state import init_state
from utils.session_restore import restore_session
from utils.persistence import (
    list_runs,
    list_artifacts,
    get_artifact_bytes,
    delete_run,
    rename_run,          # ✅ NEW
    XLSX_MIME,
)

st.set_page_config(page_title="Sessions", layout="wide")
init_state()
inject_css()
hero("Sessions")


# =========================================================
# Helpers
# =========================================================
PREFERRED_BRANDS = ["TAS", "Peugeot", "Citroen"]

def _find_artifact(run_id: int, scope: str, exact_name: str):
    for a in list_artifacts(run_id):
        if a["scope"] == scope and a["name"] == exact_name:
            return a
    return None

def _find_artifact_suffix(run_id: int, scope: str, suffix: str):
    for a in list_artifacts(run_id):
        if a["scope"] == scope and a["name"].lower().endswith(suffix.lower()):
            return a
    return None

def _download_btn(run_id: int, scope: str, exact_name: str, label: str, file_name: str | None = None):
    a = _find_artifact(run_id, scope, exact_name)
    if not a:
        st.info(f"Not found: {scope}/{exact_name}")
        return False

    b = get_artifact_bytes(a["id"])
    if not b:
        st.info(f"Empty artifact: {exact_name}")
        return False

    st.download_button(
        label=label,
        data=b,
        file_name=file_name or a["name"],
        mime=a["content_type"],
        use_container_width=True,
        key=f"dl_{run_id}_{scope}_{a['id']}",
    )
    return True

def _download_btn_suffix(run_id: int, scope: str, suffix: str, label: str, fallback_name: str):
    a = _find_artifact_suffix(run_id, scope, suffix)
    if not a:
        st.info(f"Not found: {scope}/*{suffix}")
        return False

    b = get_artifact_bytes(a["id"])
    if not b:
        st.info(f"Empty artifact: {a['name']}")
        return False

    st.download_button(
        label=label,
        data=b,
        file_name=fallback_name,
        mime=a["content_type"],
        use_container_width=True,
        key=f"dl_{run_id}_{scope}_{a['id']}",
    )
    return True

def _get_artifact_bytes_by_name(run_id: int, scope: str, exact_name: str) -> bytes | None:
    a = _find_artifact(run_id, scope, exact_name)
    if not a:
        return None
    b = get_artifact_bytes(a["id"])
    return b if b else None


# =========================================================
# Compare helpers (Vehicle Dates vs Dataset_Complet)
# =========================================================
def _normalize_vid(x: object) -> str:
    if x is None:
        return ""
    try:
        if isinstance(x, float) and pd.isna(x):
            return ""
    except Exception:
        pass

    s = str(x).strip()
    if not s or s.lower() in {"nan", "none"}:
        return ""

    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]

    s = s.replace("\u00A0", "").replace("\xa0", "").replace(" ", "")

    m = re.search(r"17-?(\d+)", s)
    if m:
        digits = m.group(1)
        return f"17-{digits.zfill(6)[-6:]}" if digits.isdigit() else ""

    if s.isdigit():
        return f"17-{s.zfill(6)[-6:]}"

    return ""

def _pick_vehicle_col(df: pd.DataFrame, candidates) -> str:
    cols = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = str(cand).strip().lower()
        if k in cols:
            return cols[k]
    for c in df.columns:
        lc = str(c).strip().lower()
        if any(str(cc).strip().lower() in lc for cc in candidates):
            return c
    raise KeyError(f"Vehicle column not found. Columns: {list(df.columns)}")

def _compute_mismatches(ds2_bytes: bytes, dataset_bytes: bytes):
    # Vehicle Dates (earliest_per_vehicle)
    df_car = pd.read_excel(BytesIO(ds2_bytes), sheet_name="earliest_per_vehicle")
    car_col = _pick_vehicle_col(df_car, ["vehicle_raw", "vehicule", "véhicule", "vehicle"])
    df_car["VehicleNorm"] = df_car[car_col].apply(_normalize_vid)
    df_car = df_car[df_car["VehicleNorm"] != ""].copy()
    car_set = set(df_car["VehicleNorm"].unique())

    # Dataset_Complet (Global Vehicle List)
    df_glob = pd.read_excel(BytesIO(dataset_bytes), sheet_name="Global Vehicle List")
    glob_col = _pick_vehicle_col(df_glob, ["véhicule id", "vehicule id", "vehicleid", "vehicle id"])
    df_glob["VehicleNorm"] = df_glob[glob_col].apply(_normalize_vid)
    df_glob = df_glob[df_glob["VehicleNorm"] != ""].copy()
    glob_set = set(df_glob["VehicleNorm"].unique())

    in_car_not_in_glob = sorted(car_set - glob_set)
    in_glob_not_in_car = sorted(glob_set - car_set)

    df_a = pd.DataFrame({"VehicleNorm": in_car_not_in_glob})
    df_b = pd.DataFrame({"VehicleNorm": in_glob_not_in_car})

    # Context: Carburant side
    if not df_a.empty:
        keep = [car_col, "VehicleNorm"]
        for c in ["source_file", "sheet", "excel_row", "vehicle_earliest_date", "vehicle_age_years", "date_raw"]:
            if c in df_car.columns:
                keep.append(c)
        df_a = df_a.merge(df_car[keep].drop_duplicates("VehicleNorm"), on="VehicleNorm", how="left")
        df_a = df_a.rename(columns={car_col: "vehicle_raw_original"})

    # Context: Dataset side
    if not df_b.empty:
        keep = [glob_col, "VehicleNorm"]
        for c in ["Total HTVA", "TotalHTVA", "Brand(s)", "Brands"]:
            if c in df_glob.columns:
                keep.append(c)
        df_b = df_b.merge(df_glob[keep].drop_duplicates("VehicleNorm"), on="VehicleNorm", how="left")
        df_b = df_b.rename(columns={glob_col: "vehicle_list_original"})

    summary = pd.DataFrame(
        {
            "Metric": [
                "Unique vehicles in carburant (normalized)",
                "Unique vehicles in Dataset_Complet Global Vehicle List (normalized)",
                "In carburant but NOT in Dataset_Complet",
                "In Dataset_Complet but NOT in carburant",
            ],
            "Value": [len(car_set), len(glob_set), len(in_car_not_in_glob), len(in_glob_not_in_car)],
        }
    )

    return df_a, df_b, summary


# =========================================================
# Pretty one-sheet Excel builder (always visible)
# =========================================================
_THIN = Side(style="thin", color="D0D7DE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_COLOR_TITLE = "1F4E79"
_COLOR_SECTION = "2F5597"
_COLOR_TABLE_HDR = "D9E1F2"
_COLOR_ZEBRA = "F7F9FC"

_TITLE_FONT = Font(bold=True, color="FFFFFF", size=13)
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
_HDR_FONT = Font(bold=True, color="0F172A")
_BODY_FONT = Font(color="0F172A")

_ALIGN_TITLE = Alignment(horizontal="left", vertical="center")
_ALIGN_HDR = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_BODY = Alignment(horizontal="left", vertical="top", wrap_text=True)

def _set_col_widths(ws, max_col: int, widths: dict[int, int]):
    for c in range(1, max_col + 1):
        w = widths.get(c, 12)
        ws.column_dimensions[get_column_letter(c)].width = min(max(w, 10), 55)

def _apply_border(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = _BORDER

def _write_main_title(ws, row: int, title: str, width_cols: int) -> int:
    fill = PatternFill("solid", fgColor=_COLOR_TITLE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = fill
    cell.font = _TITLE_FONT
    cell.alignment = _ALIGN_TITLE
    ws.row_dimensions[row].height = 26
    for c in range(1, width_cols + 1):
        cc = ws.cell(row=row, column=c)
        cc.fill = fill
        cc.border = _BORDER
    return row + 1

def _write_subtitle(ws, row: int, text: str, width_cols: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(color="334155", italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    return row + 2

def _write_section_bar(ws, row: int, title: str, width_cols: int) -> int:
    fill = PatternFill("solid", fgColor=_COLOR_SECTION)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = fill
    cell.font = _SECTION_FONT
    cell.alignment = _ALIGN_TITLE
    ws.row_dimensions[row].height = 22
    for c in range(1, width_cols + 1):
        cc = ws.cell(row=row, column=c)
        cc.fill = fill
        cc.border = _BORDER
    return row + 1

def _write_table(ws, df: pd.DataFrame, start_row: int, *, money_cols=None, int_cols=None):
    money_cols = set(money_cols or [])
    int_cols = set(int_cols or [])

    if df is None or df.empty:
        ws.cell(start_row, 1, "(no rows)").font = Font(italic=True, color="64748B")
        return start_row + 2, 1, {1: 12}

    ncols = len(df.columns)
    col_widths: dict[int, int] = {}

    header_fill = PatternFill("solid", fgColor=_COLOR_TABLE_HDR)
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(start_row, j, str(col))
        cell.fill = header_fill
        cell.font = _HDR_FONT
        cell.alignment = _ALIGN_HDR
        cell.border = _BORDER
        col_widths[j] = max(col_widths.get(j, 0), len(str(col)))

    zebra_fill = PatternFill("solid", fgColor=_COLOR_ZEBRA)
    r = start_row + 1
    for i, row_vals in enumerate(df.itertuples(index=False, name=None), start=0):
        is_zebra = (i % 2 == 1)
        for j, v in enumerate(row_vals, start=1):
            val = "" if v is None or (isinstance(v, float) and pd.isna(v)) else v
            cell = ws.cell(r, j, val)
            cell.font = _BODY_FONT
            cell.alignment = _ALIGN_BODY
            cell.border = _BORDER
            if is_zebra:
                cell.fill = zebra_fill

            colname = str(df.columns[j - 1])
            if colname in money_cols:
                cell.number_format = "0.000"
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            elif colname in int_cols:
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

            col_widths[j] = max(col_widths.get(j, 0), len(str(val)) if val is not None else 0)
        r += 1

    end_row = r - 1
    _apply_border(ws, start_row, end_row, 1, ncols)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(ncols)}{end_row}"
    return end_row + 2, ncols, col_widths

def _build_one_sheet_report_bytes(summary_df: pd.DataFrame, df_a: pd.DataFrame, df_b: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    WIDTH_COLS = 10
    r = 1

    r = _write_main_title(ws, r, "Vehicle ID Mismatch Report", WIDTH_COLS)
    r = _write_subtitle(ws, r, f"Generated at: {datetime.now():%Y-%m-%d %H:%M:%S}", WIDTH_COLS)

    r = _write_section_bar(ws, r, "SUMMARY", WIDTH_COLS)
    r, max_col, widths = _write_table(ws, summary_df, r, int_cols={"Value"})
    _set_col_widths(ws, max_col, widths)

    r = _write_section_bar(ws, r, "CARBURANT NOT IN DATASET", WIDTH_COLS)
    r, max_col, widths = _write_table(ws, df_a, r, int_cols={"excel_row", "vehicle_age_years"})
    _set_col_widths(ws, max_col, widths)

    r = _write_section_bar(ws, r, "DATASET NOT IN CARBURANT", WIDTH_COLS)
    r, max_col, widths = _write_table(ws, df_b, r, money_cols={"Total HTVA", "TotalHTVA"})
    _set_col_widths(ws, max_col, widths)

    ws.freeze_panes = "A5"
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================================================
# 1) Session selector
# =========================================================
runs = list_runs(limit=200)
if not runs:
    st.info("No sessions found yet. Run Cleaning first.")
    st.stop()

def _run_label(r: dict) -> str:
    nm = (r.get("run_name") or "").strip()
    if nm:
        return f"{nm} — Session {r['id']} — {r['created_at']}"
    return f"Session {r['id']} — {r['created_at']}"

options = {_run_label(r): r["id"] for r in runs}
labels = ["— Select a session —"] + list(options.keys())

chosen_label = st.selectbox("Select a session", labels, index=0, key="session_select")
if chosen_label == "— Select a session —":
    st.stop()

run_id = options[chosen_label]

# Hydrate session state (dashboards)
restore_session(run_id)
st.session_state["active_run_id"] = run_id


# =========================================================
# 2) Tabs
# =========================================================
st.divider()
tab1, tab2, tab3, tab4 = st.tabs(["Cleaned Files", "Global Merge", "Vehicle Dates Extraction", "Dashboards"])

with tab1:
    card_open()
    st.subheader("Cleaned files")
    cols = st.columns(3)
    for i, brand in enumerate(PREFERRED_BRANDS):
        with cols[i]:
            st.markdown(f"**{brand}**")
            _download_btn_suffix(
                run_id=run_id,
                scope=brand,
                suffix="_cleaned.xlsx",
                label=f"Download {brand}_cleaned.xlsx",
                fallback_name=f"{brand}_cleaned.xlsx",
            )
    card_close()

with tab2:
    card_open()
    st.subheader("Global merged workbook")
    _download_btn(
        run_id=run_id,
        scope="GLOBAL",
        exact_name="Dataset_Complet.xlsx",
        label="Download Dataset_Complet.xlsx",
        file_name="Dataset_Complet.xlsx",
    )
    st.caption("If it is missing, build the global merge in Results page.")
    card_close()

with tab3:
    card_open()
    st.subheader("Vehicle dates outputs")

    c1, c2 = st.columns(2)
    with c1:
        _download_btn(
            run_id=run_id,
            scope="VEHICLE_DATES",
            exact_name="vehicle_dates_all_rows.xlsx",
            label="Download all extracted dates",
            file_name="vehicle_dates_all_rows.xlsx",
        )
    with c2:
        _download_btn(
            run_id=run_id,
            scope="VEHICLE_DATES",
            exact_name="vehicle_dates_earliest_per_vehicle.xlsx",
            label="Download earliest per vehicle",
            file_name="vehicle_dates_earliest_per_vehicle.xlsx",
        )

    st.divider()
    st.subheader("Compare carburant vs Dataset_Complet")

    # ✅ Use DB artifacts directly (no re-upload needed)
    ds2_bytes = _get_artifact_bytes_by_name(run_id, "VEHICLE_DATES", "vehicle_dates_earliest_per_vehicle.xlsx")
    dataset_bytes = _get_artifact_bytes_by_name(run_id, "GLOBAL", "Dataset_Complet.xlsx")

    # Only show upload IF missing
    uploaded_dataset = None
    if not dataset_bytes:
        uploaded_dataset = st.file_uploader(
            "Dataset_Complet is missing in this session. Upload it here (.xlsx)",
            type=["xlsx"],
            key="sess_compare_upload_dataset",
        )
        if uploaded_dataset is not None:
            dataset_bytes = uploaded_dataset.getbuffer().tobytes()

    disabled = (not ds2_bytes) or (not dataset_bytes)

    if st.button("Run comparison", type="primary", disabled=disabled, use_container_width=True, key="sess_run_compare"):
        try:
            with st.spinner("Comparing vehicle IDs..."):
                df_a, df_b, summary = _compute_mismatches(ds2_bytes, dataset_bytes)
            st.session_state["sess_compare_summary"] = summary
            st.session_state["sess_compare_a"] = df_a
            st.session_state["sess_compare_b"] = df_b
            st.success("✅ Comparison complete.")
            st.rerun()
        except Exception as e:
            st.error(f"Comparison failed: {e}")
            st.exception(e)

    summary = st.session_state.get("sess_compare_summary")
    df_a = st.session_state.get("sess_compare_a")
    df_b = st.session_state.get("sess_compare_b")

    if isinstance(summary, pd.DataFrame) and isinstance(df_a, pd.DataFrame) and isinstance(df_b, pd.DataFrame):
        st.dataframe(summary, use_container_width=True, height=170)

        left, right = st.columns(2)
        with left:
            st.markdown("**In carburant (vehicle_dates) but NOT in Dataset_Complet**")
            st.dataframe(df_a, use_container_width=True, height=360)
        with right:
            st.markdown("**In Dataset_Complet but NOT in carburant (vehicle_dates)**")
            st.dataframe(df_b, use_container_width=True, height=360)

        report_payload = _build_one_sheet_report_bytes(summary, df_a, df_b)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            "⬇️ Download mismatch report (Excel)",
            data=report_payload,
            file_name=f"vehicle_id_mismatch_report_{ts}.xlsx",
            mime=XLSX_MIME,
            use_container_width=True,
            key=f"sess_dl_mismatch_{ts}",
        )
    else:
        if not ds2_bytes:
            st.info("This session has no vehicle_dates_earliest_per_vehicle.xlsx yet. Run Vehicle Dates Extraction first.")
        elif not dataset_bytes:
            st.info("This session has no Dataset_Complet.xlsx saved yet. Build Global Merge first, or upload it above.")

    card_close()

with tab4:
    card_open()
    st.subheader("Dashboards")

    st.session_state.setdefault("dash_inline_loaded", False)
    cA, cB = st.columns([1, 1])
    with cA:
        if st.button("Load dashboards", type="primary", use_container_width=True):
            st.session_state["dash_inline_loaded"] = True
    with cB:
        if st.button("Hide dashboards", use_container_width=True):
            st.session_state["dash_inline_loaded"] = False

    if st.session_state.get("dash_inline_loaded"):
        st.divider()
        from utils.dashboards_core import render_dashboards
        render_dashboards()
    else:
        st.caption("Click **Load dashboards** to render the dashboards here inside Sessions.")

    card_close()


# =========================================================
# Session options (Rename + Delete)
# =========================================================
st.divider()
with st.expander("Session options"):
    # --- Rename session ---
    st.markdown("### Rename session")
    current_run = next((r for r in runs if r["id"] == run_id), None)
    current_name = (current_run.get("run_name") if current_run else "") or ""

    new_name = st.text_input(
        "Session name",
        value=current_name,
        placeholder="e.g., Feb 2026 import – Peugeot/Citroen",
        key=f"sess_rename_{run_id}",
    )

    c1, c2 = st.columns([1, 2])
    with c1:
        if st.button("Save name", use_container_width=True, key=f"sess_rename_save_{run_id}"):
            try:
                rename_run(run_id, new_name)
                st.success("Session renamed.")
                st.rerun()
            except Exception as e:
                st.error(f"Rename failed: {e}")

    st.divider()

    # --- Delete session ---
    confirm = st.checkbox("I understand this will permanently delete the session and all saved files.")
    if st.button("Delete this session", type="primary", disabled=not confirm):
        delete_run(run_id)
        st.session_state.pop("active_run_id", None)
        st.session_state.pop("results", None)
        st.session_state.pop("sess_compare_summary", None)
        st.session_state.pop("sess_compare_a", None)
        st.session_state.pop("sess_compare_b", None)
        st.success("Session deleted.")
        st.rerun()