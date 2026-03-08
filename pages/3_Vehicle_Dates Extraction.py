# pages/3_Vehicle_Dates Extraction.py
import os
import tempfile
import re
from io import BytesIO
from datetime import datetime

import streamlit as st
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.ui import inject_css, hero, section, card_open, card_close
from utils.state import init_state
from utils.vehicle_dates import run_vehicle_date_extraction
from utils.persistence import save_artifact, save_df_parquet, XLSX_MIME

# ✅ MUST BE FIRST Streamlit call
st.set_page_config(page_title="Vehicle Dates", layout="wide")

init_state()
inject_css()
hero("Extract Vehicle Dates")


# -------------------------
# Reset only this feature
# -------------------------
colA, colB = st.columns([1, 6])
with colA:
    if st.button("Reset"):
        st.session_state["vehicle_dates_ds1"] = None
        st.session_state["vehicle_dates_ds2"] = None
        st.session_state["df_vehicle_dates_all"] = None
        st.session_state["df_vehicle_dates_earliest"] = None
        st.session_state["df_vehicle_dates_oldest"] = None

        # comparison outputs
        st.session_state.pop("compare_summary_full", None)
        st.session_state.pop("compare_df_car_not_in_ds", None)
        st.session_state.pop("compare_df_ds_not_in_car", None)
        st.session_state.pop("compare_df_car_all", None)
        st.session_state.pop("compare_df_ds_all", None)

        st.rerun()


# -------------------------
# Upload
# -------------------------
uploaded = st.file_uploader(
    "Select ALL Excel files here (You can upload 40 files in one go.)",
    type=["xlsx"],
    accept_multiple_files=True,
    key="veh_dates_files",
)

st.divider()


# -------------------------
# Run button (only once)
# -------------------------
already_done = st.session_state.get("vehicle_dates_ds2") is not None
run_disabled = (not uploaded) or already_done

if st.button("Run Vehicle Dates Extraction", type="primary", disabled=run_disabled):
    with st.spinner("Processing files..."):
        tmp_dir = tempfile.mkdtemp(prefix="veh_dates_")
        input_paths = []

        for uf in uploaded:
            p = os.path.join(tmp_dir, uf.name)
            with open(p, "wb") as f:
                f.write(uf.getbuffer())
            input_paths.append(p)

        ds1_bytes, ds2_bytes = run_vehicle_date_extraction(input_paths)

    # Store in session
    st.session_state["vehicle_dates_ds1"] = ds1_bytes
    st.session_state["vehicle_dates_ds2"] = ds2_bytes

    # Optional previews saved for later (dashboard)
    df2 = None
    try:
        df1 = pd.read_excel(BytesIO(ds1_bytes), sheet_name="extraction")
        df2 = pd.read_excel(BytesIO(ds2_bytes), sheet_name="earliest_per_vehicle")

        st.session_state["df_vehicle_dates_all"] = df1
        st.session_state["df_vehicle_dates_earliest"] = df2
        st.session_state["df_vehicle_dates_oldest"] = df2  # dashboard compat
    except Exception:
        pass

    # Persist into SAME session (if exists)
    run_id = st.session_state.get("active_run_id")
    if run_id:
        try:
            save_artifact(run_id, "VEHICLE_DATES", "vehicle_dates_all_rows.xlsx", XLSX_MIME, ds1_bytes)
            save_artifact(run_id, "VEHICLE_DATES", "vehicle_dates_earliest_per_vehicle.xlsx", XLSX_MIME, ds2_bytes)

            if isinstance(df2, pd.DataFrame) and not df2.empty:
                save_df_parquet(run_id, "VEHICLE_DATES", "vehicle_dates_earliest.parquet", df2)
        except Exception as e:
            st.warning(f"Session save failed (vehicle dates): {e}")

    st.success("✅ Done. Download your results below.")
    st.rerun()


# -------------------------
# Download section
# -------------------------
ds1 = st.session_state.get("vehicle_dates_ds1")
ds2 = st.session_state.get("vehicle_dates_ds2")

if ds1 and ds2:
    section("Vehicle dates outputs")
    card_open()

    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "Download all extracted dates",
            data=ds1,
            file_name="vehicle_dates_all_rows.xlsx",
            mime=XLSX_MIME,
            use_container_width=True,
            key="dl_vehicle_dates_all",
        )
    with c2:
        st.download_button(
            "Download earliest per vehicle",
            data=ds2,
            file_name="vehicle_dates_earliest_per_vehicle.xlsx",
            mime=XLSX_MIME,
            use_container_width=True,
            key="dl_vehicle_dates_earliest",
        )

    with st.expander("Preview"):
        df_all = st.session_state.get("df_vehicle_dates_all")
        df_old = st.session_state.get("df_vehicle_dates_oldest")
        if isinstance(df_all, pd.DataFrame) and not df_all.empty:
            st.subheader("All dates extracted (preview)")
            st.dataframe(df_all.head(30), use_container_width=True)
        if isinstance(df_old, pd.DataFrame) and not df_old.empty:
            st.subheader("Earliest per vehicle (preview)")
            st.dataframe(df_old.head(30), use_container_width=True)

    card_close()
else:
    st.info("Upload your files and click **Run Vehicle Dates Extraction**.")


# =========================================================
# ✅ Compare Carburant vs Dataset_Complet (Global Vehicle List)
# =========================================================
st.divider()
section("Compare Carburant vs Dataset_Complet")
card_open()

st.caption("Compares vehicle_dates_earliest_per_vehicle (vehicle_raw) vs Dataset_Complet.xlsx → Global Vehicle List.")


def _normalize_vid(x: object):
    """Return (norm, status). norm is '17-XXXXXX' or pd.NA."""
    if x is None:
        return (pd.NA, "missing")

    try:
        if isinstance(x, float) and pd.isna(x):
            return (pd.NA, "missing")
    except Exception:
        pass

    s = str(x).strip()
    if not s or s.lower() in {"nan", "none"}:
        return (pd.NA, "missing")

    # "12345.0" -> "12345"
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]

    # remove spaces + NBSP
    s = s.replace("\u00A0", "").replace("\xa0", "")
    s = re.sub(r"\s+", "", s)

    # prefer explicit 17 prefix patterns
    m = re.search(r"(?:^|[^0-9])17[-_./]?(?P<num>\d+)(?:[^0-9]|$)", s)
    if m:
        digits = m.group("num")
        return (f"17-{digits.zfill(6)[-6:]}", "ok:has_17_prefix")

    # if just digits
    if s.isdigit():
        return (f"17-{s.zfill(6)[-6:]}", "ok:digits_only")

    # salvage any digits (so you can still compare + inspect)
    d = re.findall(r"\d+", s)
    if d:
        digits = "".join(d)
        return (f"17-{digits.zfill(6)[-6:]}", "warn:salvaged_digits")

    return (pd.NA, "invalid:no_digits")


def _pick_vehicle_col(df: pd.DataFrame, candidates) -> str:
    cols = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = str(cand).strip().lower()
        if k in cols:
            return cols[k]
    for c in df.columns:
        lc = str(c).strip().lower()
        if any(str(cc).strip().lower() in lc for cc in candidates):
            return c
    raise KeyError(f"Vehicle column not found. Columns: {list(df.columns)}")


def _clean_id(x: object) -> str:
    """Minimal cleaning only: stringify + strip. No normalization."""
    if x is None:
        return ""
    try:
        if isinstance(x, float) and pd.isna(x):
            return ""
    except Exception:
        pass
    s = str(x).strip()
    if not s or s.lower() in {"nan", "none"}:
        return ""
    # optional: remove NBSP that Excel sometimes inserts
    s = s.replace("\u00A0", "").replace("\xa0", "").strip()
    return s


def _compute_mismatches(vehicle_dates_ds2_bytes: bytes, dataset_complet_bytes: bytes):
    # --- Carburant list (earliest_per_vehicle) ---
    df_car = pd.read_excel(BytesIO(vehicle_dates_ds2_bytes), sheet_name="earliest_per_vehicle")
    car_col = _pick_vehicle_col(df_car, ["vehicle_raw", "vehicule", "véhicule", "vehicle"])
    df_car["VehicleID"] = df_car[car_col].apply(_clean_id)

    # Keep all (for display), but build set using non-empty IDs
    car_all = df_car[[car_col, "VehicleID"]].copy().rename(columns={car_col: "vehicle_raw_original"})
    car_set = set(df_car.loc[df_car["VehicleID"] != "", "VehicleID"].unique())

    # --- Dataset_Complet list (Global Vehicle List) ---
    df_glob = pd.read_excel(BytesIO(dataset_complet_bytes), sheet_name="Global Vehicle List")
    glob_col = _pick_vehicle_col(df_glob, ["véhicule id", "vehicule id", "vehicleid", "vehicle id"])
    df_glob["VehicleID"] = df_glob[glob_col].apply(_clean_id)

    ds_all = df_glob[[glob_col, "VehicleID"]].copy().rename(columns={glob_col: "vehicle_list_original"})
    glob_set = set(df_glob.loc[df_glob["VehicleID"] != "", "VehicleID"].unique())

    # --- Differences (exact string match) ---
    in_car_not_in_glob = sorted(car_set - glob_set)
    in_glob_not_in_car = sorted(glob_set - car_set)

    df_a = pd.DataFrame({"VehicleID": in_car_not_in_glob})
    df_b = pd.DataFrame({"VehicleID": in_glob_not_in_car})

    # Add context (carburant)
    if not df_a.empty:
        keep = [car_col, "VehicleID"]
        for c in ["source_file", "sheet", "excel_row", "vehicle_earliest_date", "vehicle_age_years", "date_raw"]:
            if c in df_car.columns:
                keep.append(c)
        df_a = df_a.merge(df_car[keep].drop_duplicates("VehicleID"), on="VehicleID", how="left")
        df_a = df_a.rename(columns={car_col: "vehicle_raw_original"})

    # Add context (dataset)
    if not df_b.empty:
        keep = [glob_col, "VehicleID"]
        for c in ["Total HTVA", "TotalHTVA", "Brand(s)", "Brands"]:
            if c in df_glob.columns:
                keep.append(c)
        df_b = df_b.merge(df_glob[keep].drop_duplicates("VehicleID"), on="VehicleID", how="left")
        df_b = df_b.rename(columns={glob_col: "vehicle_list_original"})

    summary_full = pd.DataFrame(
        {
            "Metric": [
                "Unique vehicles in carburant (raw)",
                "Unique vehicles in Dataset_Complet Global Vehicle List (raw)",
                "In carburant but NOT in Dataset_Complet",
                "In Dataset_Complet but NOT in carburant",
                "Empty/blank vehicle IDs in carburant (rows)",
                "Empty/blank vehicle IDs in dataset (rows)",
            ],
            "Value": [
                len(car_set),
                len(glob_set),
                len(in_car_not_in_glob),
                len(in_glob_not_in_car),
                int((df_car["VehicleID"] == "").sum()),
                int((df_glob["VehicleID"] == "").sum()),
            ],
        }
    )

    return df_a, df_b, summary_full, car_all, ds_all


# -------------------------
# Pretty ONE-sheet Excel report builder
# -------------------------
_THIN = Side(style="thin", color="D0D7DE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Colors
_COLOR_TITLE = "1F4E79"      # dark blue
_COLOR_SECTION = "2F5597"    # section blue
_COLOR_TABLE_HDR = "D9E1F2"  # table header light
_COLOR_ZEBRA = "F7F9FC"      # zebra stripe

_TITLE_FONT = Font(bold=True, color="FFFFFF", size=13)
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
_HDR_FONT = Font(bold=True, color="0F172A")
_BODY_FONT = Font(color="0F172A")

_ALIGN_TITLE = Alignment(horizontal="left", vertical="center")
_ALIGN_HDR = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_BODY = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _set_col_widths(ws, max_col: int, widths: dict[int, int]):
    for c in range(1, max_col + 1):
        w = widths.get(c, 12)
        ws.column_dimensions[get_column_letter(c)].width = min(max(w, 10), 55)


def _apply_border(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = _BORDER


def _write_main_title(ws, row: int, title: str, width_cols: int) -> int:
    fill = PatternFill("solid", fgColor=_COLOR_TITLE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = fill
    cell.font = _TITLE_FONT
    cell.alignment = _ALIGN_TITLE
    ws.row_dimensions[row].height = 26
    for c in range(1, width_cols + 1):
        cc = ws.cell(row=row, column=c)
        cc.fill = fill
        cc.border = _BORDER
    return row + 1


def _write_subtitle(ws, row: int, text: str, width_cols: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(color="334155", italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    return row + 2


def _write_section_bar(ws, row: int, title: str, width_cols: int) -> int:
    fill = PatternFill("solid", fgColor=_COLOR_SECTION)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = fill
    cell.font = _SECTION_FONT
    cell.alignment = _ALIGN_TITLE
    ws.row_dimensions[row].height = 22
    for c in range(1, width_cols + 1):
        cc = ws.cell(row=row, column=c)
        cc.fill = fill
        cc.border = _BORDER
    return row + 1


def _write_table(ws, df: pd.DataFrame, start_row: int, *, money_cols=None, int_cols=None):
    money_cols = set(money_cols or [])
    int_cols = set(int_cols or [])

    if df is None or df.empty:
        ws.cell(start_row, 1, "(no rows)").font = Font(italic=True, color="64748B")
        return start_row + 2, 1, {1: 12}

    ncols = len(df.columns)
    col_widths: dict[int, int] = {}

    # Header row
    header_fill = PatternFill("solid", fgColor=_COLOR_TABLE_HDR)
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(start_row, j, str(col))
        cell.fill = header_fill
        cell.font = _HDR_FONT
        cell.alignment = _ALIGN_HDR
        cell.border = _BORDER
        col_widths[j] = max(col_widths.get(j, 0), len(str(col)))

    # Data rows with zebra stripes
    zebra_fill = PatternFill("solid", fgColor=_COLOR_ZEBRA)

    r = start_row + 1
    for i, row_vals in enumerate(df.itertuples(index=False, name=None), start=0):
        is_zebra = (i % 2 == 1)
        for j, v in enumerate(row_vals, start=1):
            val = "" if v is None or (isinstance(v, float) and pd.isna(v)) else v
            cell = ws.cell(r, j, val)
            cell.font = _BODY_FONT
            cell.alignment = _ALIGN_BODY
            cell.border = _BORDER
            if is_zebra:
                cell.fill = zebra_fill

            colname = str(df.columns[j - 1])
            if colname in money_cols:
                cell.number_format = "0.000"
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            elif colname in int_cols:
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

            col_widths[j] = max(col_widths.get(j, 0), len(str(val)) if val is not None else 0)

        r += 1

    end_row = r - 1
    _apply_border(ws, start_row, end_row, 1, ncols)

    # Apply autofilter to the table
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(ncols)}{end_row}"

    return end_row + 2, ncols, col_widths


def _build_one_sheet_report_bytes(summary_df: pd.DataFrame, df_a: pd.DataFrame, df_b: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    WIDTH_COLS = 10  # fixed visual width across the sheet
    r = 1

    r = _write_main_title(ws, r, "Vehicle ID Mismatch Report", WIDTH_COLS)
    r = _write_subtitle(ws, r, f"Generated at: {datetime.now():%Y-%m-%d %H:%M:%S}", WIDTH_COLS)

    # Summary
    r = _write_section_bar(ws, r, "SUMMARY", WIDTH_COLS)
    r, max_col, widths = _write_table(ws, summary_df, r, int_cols={"Value"})
    _set_col_widths(ws, max_col, widths)

    # Carburant not in dataset
    r = _write_section_bar(ws, r, "CARBURANT NOT IN DATASET", WIDTH_COLS)
    r, max_col, widths = _write_table(ws, df_a, r, int_cols={"excel_row", "vehicle_age_years"})
    _set_col_widths(ws, max_col, widths)

    # Dataset not in carburant
    r = _write_section_bar(ws, r, "DATASET NOT IN CARBURANT", WIDTH_COLS)
    r, max_col, widths = _write_table(ws, df_b, r, money_cols={"Total HTVA", "TotalHTVA"})
    _set_col_widths(ws, max_col, widths)

    # Freeze after title+subtitle
    ws.freeze_panes = "A5"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# Inputs for comparison
ds2_bytes = st.session_state.get("vehicle_dates_ds2")
dataset_bytes = st.session_state.get("global_merge_bytes")

uploaded_dataset = st.file_uploader(
    "If Dataset_Complet is not in this session, upload it here (.xlsx)",
    type=["xlsx"],
    key="upload_dataset_complet_compare",
)
if uploaded_dataset is not None:
    dataset_bytes = uploaded_dataset.getbuffer().tobytes()

run_compare_disabled = (not ds2_bytes) or (not dataset_bytes)

if st.button("Run comparison", type="primary", disabled=run_compare_disabled, use_container_width=True):
    try:
        with st.spinner("Comparing vehicle IDs..."):
            df_car_missing, df_ds_missing, summary_full, df_car_all, df_ds_all = _compute_mismatches(ds2_bytes, dataset_bytes)

        st.session_state["compare_df_car_not_in_ds"] = df_car_missing
        st.session_state["compare_df_ds_not_in_car"] = df_ds_missing
        st.session_state["compare_summary_full"] = summary_full
        st.session_state["compare_df_car_all"] = df_car_all
        st.session_state["compare_df_ds_all"] = df_ds_all

        st.success("✅ Comparison complete.")
        st.rerun()
    except Exception as e:
        st.error(f"Comparison failed: {e}")
        st.exception(e)

df_car_missing = st.session_state.get("compare_df_car_not_in_ds")
df_ds_missing = st.session_state.get("compare_df_ds_not_in_car")
summary_full = st.session_state.get("compare_summary_full")
df_car_all = st.session_state.get("compare_df_car_all")
df_ds_all = st.session_state.get("compare_df_ds_all")

if isinstance(summary_full, pd.DataFrame) and isinstance(df_car_missing, pd.DataFrame) and isinstance(df_ds_missing, pd.DataFrame):
    st.dataframe(summary_full, use_container_width=True, height=240)

    with st.expander("Show ALL vehicles (raw → normalized)"):
        if isinstance(df_car_all, pd.DataFrame):
            st.markdown("#### All vehicles from carburant")
            st.dataframe(df_car_all, use_container_width=True, height=260)

        if isinstance(df_ds_all, pd.DataFrame):
            st.markdown("#### All vehicles from Dataset_Complet")
            st.dataframe(df_ds_all, use_container_width=True, height=260)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**In carburant (vehicle_dates) but NOT in Dataset_Complet**")
        st.dataframe(df_car_missing, use_container_width=True, height=360)
    with c2:
        st.markdown("**In Dataset_Complet but NOT in carburant (vehicle_dates)**")
        st.dataframe(df_ds_missing, use_container_width=True, height=360)

    # ✅ Pretty ONE-sheet report
    report_payload = _build_one_sheet_report_bytes(summary_full, df_car_missing, df_ds_missing)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    st.download_button(
        "⬇️ Download mismatch report (Excel)",
        data=report_payload,
        file_name=f"vehicle_id_mismatch_report_{ts}.xlsx",
        mime=XLSX_MIME,
        use_container_width=True,
        key=f"dl_vehicle_mismatch_report_{ts}",
    )
else:
    if not ds2_bytes:
        st.info("Run Vehicle Dates Extraction first (so earliest_per_vehicle exists).")
    elif not dataset_bytes:
        st.info("Build Global Merge first (Dataset_Complet) OR upload Dataset_Complet.xlsx above.")

card_close()